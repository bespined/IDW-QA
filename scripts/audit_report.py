#!/usr/bin/env python3
"""Audit Report Generator — produce a polished, shareable HTML report from audit results.

Generates a comprehensive HTML report covering:
  - Design Standards (25 ASU standards) — Met / Partially Met / Not Met
  - QA Categories (19 structural checks) — Pass / Warn / Fail
  - Accessibility (WCAG 2.1 AA) — Critical / Warning / Info
  - Course Readiness (9 categories) — Pass / Fail

Usage:
    python audit_report.py --input audit_results.json         # From saved audit JSON
    python audit_report.py --input audit_results.json --open  # Generate and open
    python audit_report.py --demo                              # Generate demo report with sample data

Reports are saved directly to reports/{COURSE-CODE_TERM}/ with timestamped filenames.
Use --open to view in the default browser after generation.
"""

import argparse
import json
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path

# Logging
try:
    from idw_logger import get_logger
    _log = get_logger("audit_report")
except ImportError:
    import logging
    _log = logging.getLogger("audit_report")

try:
    from idw_metrics import track as _track
except ImportError:
    def _track(*a, **k): pass


PLUGIN_ROOT = Path(__file__).resolve().parents[1]
STAGING_DIR = PLUGIN_ROOT / "staging"
REPORTS_DIR = PLUGIN_ROOT / "reports"


def _resolve_auditor(data: dict) -> str:
    """Resolve auditor name: env var > tester lookup > data dict > fallback."""
    if os.getenv("IDW_AUDITOR_NAME"):
        return os.getenv("IDW_AUDITOR_NAME")
    # Try looking up tester name from IDW_TESTER_ID
    tester_id = os.getenv("IDW_TESTER_ID", "").strip()
    if tester_id:
        try:
            from role_gate import get_current_tester
            tester = get_current_tester()
            if tester and tester.get("name"):
                return tester["name"]
        except Exception:
            pass
    return data.get("auditor") or "ID Workbench"


# ── RLHF Supabase Integration ──────────────────────────────────────
def _get_supabase_config():
    """Load Supabase credentials from .env / .env.local."""
    url = os.getenv("SUPABASE_URL", "")
    service_key = os.getenv("SUPABASE_SERVICE_KEY", "")
    return url, service_key


def _supabase_post(url, key, table, rows):
    """POST rows to a Supabase table. Returns inserted rows or None."""
    import requests
    resp = requests.post(
        f"{url}/rest/v1/{table}",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        },
        json=rows,
        timeout=30,
    )
    if resp.status_code in (200, 201):
        return resp.json()
    _log.warning("Supabase POST to %s failed: %s %s", table, resp.status_code, resp.text[:200])
    return None


def _supabase_patch(url, key, table, row_id, updates):
    """PATCH a single row in a Supabase table by id."""
    import requests
    resp = requests.patch(
        f"{url}/rest/v1/{table}?id=eq.{row_id}",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
        },
        json=updates,
        timeout=15,
    )
    if resp.status_code not in (200, 204):
        _log.warning("Supabase PATCH %s/%s failed: %s", table, row_id, resp.status_code)


def _supabase_upload_file(url, key, bucket, path_in_bucket, local_path):
    """Upload a file to Supabase Storage. Returns public URL or None."""
    import requests
    content_type = "text/html" if local_path.endswith(".html") else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    with open(local_path, "rb") as f:
        resp = requests.post(
            f"{url}/storage/v1/object/{bucket}/{path_in_bucket}",
            headers={
                "apikey": key,
                "Authorization": f"Bearer {key}",
                "Content-Type": content_type,
                "x-upsert": "true",
            },
            data=f,
            timeout=60,
        )
    if resp.status_code in (200, 201):
        return f"{url}/storage/v1/object/public/{bucket}/{path_in_bucket}"
    _log.warning("Supabase upload %s failed: %s %s", path_in_bucket, resp.status_code, resp.text[:200])
    return None


def push_to_rlhf(data: dict, html_path: str = None, xlsx_path: str = None):
    """Push audit findings to Supabase for the RLHF feedback loop.

    Creates an audit_session row, then inserts all findings from all sections.
    Optionally uploads HTML/XLSX report files to the rlhf-reports bucket.
    Non-blocking — failures are logged but never raise.
    """
    data = _normalize_audit_data(data)
    sb_url, sb_key = _get_supabase_config()
    if not sb_url or not sb_key:
        _log.info("Supabase not configured — skipping RLHF push")
        return None

    try:
        course = data.get("course", {})
        sections = data.get("sections", {})
        auditor = _resolve_auditor(data)

        # Compute scores for session row
        ds = sections.get("design_standards", {})
        ds_summary = ds.get("summary", {})
        ds_na = ds_summary.get("Not Auditable", 0)
        ds_total = sum(ds_summary.values()) - ds_na
        ds_met = ds_summary.get("Met", 0)
        ds_score = round(ds_met / ds_total * 100) if ds_total > 0 else 0

        qa = sections.get("qa_categories", {})
        qa_summary = qa.get("summary", {})
        qa_total = sum(qa_summary.values())
        qa_pass = qa_summary.get("Pass", 0)
        qa_score = round(qa_pass / qa_total * 100) if qa_total > 0 else 0

        a11y = sections.get("accessibility", {})
        a11y_summary = a11y.get("summary", {})
        a11y_critical = a11y_summary.get("Critical", 0)
        a11y_score = 100 if a11y_critical == 0 else max(0, 100 - a11y_critical * 25)

        readiness = sections.get("readiness", {})
        r_cats = readiness.get("categories", [])
        r_pass = sum(1 for c in r_cats if c.get("status") == "Pass")
        r_total = len(r_cats)
        r_score = round(r_pass / r_total * 100) if r_total > 0 else 0

        overall = round(ds_score * 0.4 + qa_score * 0.3 + a11y_score * 0.15 + r_score * 0.15)

        # 1. Create session
        session_row = {
            "course_id": str(course.get("id", "")),
            "course_name": course.get("name", ""),
            "course_code": course.get("course_code", course.get("name", "")),
            "term": course.get("term", ""),
            "auditor_id": auditor,
            "overall_score": data.get("overall_score", overall),
            "standards_score": data.get("readiness_score", ds_score),  # readiness = standards for display
            "a11y_score": data.get("a11y_score", a11y_score),
            "qa_score": data.get("design_score") or qa_score,  # design score if available
            "readiness_score": data.get("readiness_score", r_score),
            # Phase 2 fields
            "audit_purpose": data.get("audit_purpose", "self_audit"),
            "audit_round": data.get("audit_round", 1),
            "status": data.get("audit_status", "in_progress"),
            "plugin_version": data.get("plugin_version", "0.3.0"),
        }
        inserted = _supabase_post(sb_url, sb_key, "audit_sessions", [session_row])
        if not inserted:
            _log.warning("Failed to create audit session in Supabase")
            return None
        session_id = inserted[0]["id"]
        _log.info("RLHF session created: %s", session_id)

        # 2. Upload report files to rlhf-reports bucket
        course_code = session_row["course_code"] or "unknown"
        if html_path and os.path.exists(html_path):
            html_url = _supabase_upload_file(
                sb_url, sb_key, "rlhf-reports",
                f"{course_code}/{session_id}/report.html", html_path
            )
            if html_url:
                _supabase_patch(sb_url, sb_key, "audit_sessions", session_id,
                                {"report_html_url": html_url})
        if xlsx_path and os.path.exists(xlsx_path):
            xlsx_url = _supabase_upload_file(
                sb_url, sb_key, "rlhf-reports",
                f"{course_code}/{session_id}/report.xlsx", xlsx_path
            )
            if xlsx_url:
                _supabase_patch(sb_url, sb_key, "audit_sessions", session_id,
                                {"report_xlsx_url": xlsx_url})

        # 3. Insert findings from all sections
        findings_rows = []

        # Design standards — push per-criterion findings when available
        for item in ds.get("items", []):
            criteria_results = item.get("criteria_results", [])
            if criteria_results:
                # Push one finding per criterion (granular)
                for cr in criteria_results:
                    # Resilient field extraction — Claude may use different field names
                    # Criterion question: look in criterion_text, text, description
                    crit_question = (
                        cr.get("criterion_text")
                        or cr.get("text")
                        or cr.get("description")
                        or ""
                    )
                    # Evidence: look in evidence, detail, reasoning, content_excerpt
                    crit_evidence = (
                        cr.get("evidence")
                        or cr.get("detail")
                        or cr.get("reasoning")
                        or cr.get("content_excerpt")
                        or ""
                    )
                    # If evidence looks like a question and question is empty, swap
                    if not crit_question and crit_evidence and "?" in crit_evidence:
                        crit_question = crit_evidence
                        crit_evidence = ""
                    # If question looks like evidence (no "?" and > 20 chars), swap
                    if crit_question and "?" not in crit_question and len(crit_question) > 50 and not crit_evidence:
                        crit_evidence = crit_question
                        crit_question = ""

                    findings_rows.append({
                        "session_id": session_id,
                        "finding_type": "design",
                        "standard_id": item.get("id", ""),
                        "page_url": cr.get("page_url", item.get("page_url", "")),
                        "page_title": item.get("name", ""),  # Standard name (for grouping)
                        "ai_verdict": cr.get("status", "").lower().replace(" ", "_"),
                        "ai_reasoning": crit_question,      # What was checked
                        "content_excerpt": crit_evidence,    # What was found
                        "confidence_tier": (cr.get("confidence", item.get("confidence", ""))).lower() or None,
                        "reviewer_tier": cr.get("reviewer_tier", item.get("reviewer_tier", "id")),
                        "canvas_link": cr.get("canvas_link", item.get("canvas_link")),
                        "criterion_id": cr.get("criterion_id"),
                        "category": "crc" if item.get("id", "").startswith("crc") else "design_standard",
                        "remediation_requested": False,
                    })
            else:
                # Fallback: push standard-level finding (legacy)
                findings_rows.append({
                    "session_id": session_id,
                    "finding_type": "design",
                    "standard_id": item.get("id", ""),
                    "page_url": item.get("page_url", ""),
                    "page_title": item.get("page_title", item.get("name", "")),
                    "ai_verdict": item.get("status", "").lower().replace(" ", "_"),
                    "ai_reasoning": item.get("evidence", ""),
                    "content_excerpt": item.get("content_excerpt", item.get("recommendation", "")),
                    "confidence_tier": (item.get("confidence", "")).lower() or None,
                    "reviewer_tier": item.get("reviewer_tier", "id"),
                    "canvas_link": item.get("canvas_link"),
                    "criterion_id": item.get("criterion_id"),
                    "category": "crc" if item.get("id", "").startswith("crc") else "design_standard",
                    "remediation_requested": False,
                })

        # QA categories
        for item in qa.get("items", []):
            findings_rows.append({
                "session_id": session_id,
                "finding_type": "design",
                "standard_id": item.get("id", ""),
                "page_url": "",
                "page_title": item.get("category", item.get("name", "")),
                "ai_verdict": item.get("status", "").lower(),
                "ai_reasoning": item.get("detail", ""),
                "content_excerpt": "",
                "confidence_tier": None,
                "reviewer_tier": "id_assistant",
                "canvas_link": None,
                "criterion_id": item.get("id"),
                "category": "design_standard",
                "remediation_requested": False,
            })

        # Accessibility
        for item in a11y.get("items", []):
            findings_rows.append({
                "session_id": session_id,
                "finding_type": "accessibility",
                "standard_id": "22",
                "page_url": item.get("page_url", ""),
                "page_title": item.get("page", ""),
                "ai_verdict": item.get("severity", "").lower(),
                "ai_reasoning": item.get("issue", ""),
                "content_excerpt": item.get("element", ""),
                "confidence_tier": None,
                "reviewer_tier": "id_assistant",
                "canvas_link": item.get("canvas_link"),
                "criterion_id": None,
                "category": "design_standard",
                "remediation_requested": False,
            })

        # Readiness
        for item in r_cats:
            findings_rows.append({
                "session_id": session_id,
                "finding_type": "readiness",
                "standard_id": item.get("id", "crc"),
                "page_url": "",
                "page_title": item.get("name", ""),
                "ai_verdict": item.get("status", "").lower(),
                "ai_reasoning": item.get("detail", ""),
                "content_excerpt": "",
                "confidence_tier": None,
                "reviewer_tier": "id_assistant",
                "canvas_link": None,
                "criterion_id": item.get("criterion_id"),
                "category": "crc",
                "remediation_requested": False,
            })

        if findings_rows:
            # Supabase REST API supports batch insert
            batch_size = 50
            total_inserted = 0
            for i in range(0, len(findings_rows), batch_size):
                batch = findings_rows[i:i + batch_size]
                result = _supabase_post(sb_url, sb_key, "audit_findings", batch)
                if result:
                    total_inserted += len(result)
            _log.info("RLHF findings pushed: %d/%d", total_inserted, len(findings_rows))

        return session_id

    except Exception as e:
        _log.warning("RLHF push failed (non-blocking): %s", e)
        return None


def _get_course_info(data: dict = None):
    """Read course name and term for report naming.

    Uses the exact Canvas course name — no regex parsing. Priority:
      1. ``data`` dict — uses data['course']['name'] or data['course_name']
         If course_id is present, fetches name + term from Canvas API to fill gaps.
      2. course-config.json in plugin root
      3. Canvas API via course_id from .env
      4. Fallback: UNKNOWN / No-Term
    """
    # 1. Extract from input data if provided
    if data:
        course_block = data.get("course", {})
        course_name = (
            course_block.get("name")
            or data.get("course_name")
            or course_block.get("code")
            or course_block.get("course_code")
            or data.get("course_code")
            or ""
        )
        term = course_block.get("term") or data.get("term") or ""
        course_id = str(
            course_block.get("id")
            or data.get("course_id")
            or ""
        )

        # Fetch from Canvas API to fill in missing name or term
        if course_id and (not course_name or not term):
            api_info = _fetch_course_from_canvas(course_id)
            if not course_name:
                course_name = api_info.get("name", "")
            if not term:
                term = api_info.get("term", "")

        if course_name:
            return course_name, term or "No-Term", course_name

    # 2. Fall back to course-config.json
    config_path = PLUGIN_ROOT / "course-config.json"
    if config_path.exists():
        try:
            config = json.loads(config_path.read_text(encoding="utf-8"))
            course_title = config.get("course", {}).get("title", "")
            course_code = config.get("course", {}).get("course_code", "")
            term = config.get("course", {}).get("term", "")
            name = course_title or course_code
            if name:
                return name, term or "No-Term", course_title
        except Exception:
            pass

    # 3. Try Canvas API via .env course_id
    try:
        from dotenv import load_dotenv
        load_dotenv(PLUGIN_ROOT / ".env")
        load_dotenv(PLUGIN_ROOT / ".env.local", override=True)
        cid = os.getenv("CANVAS_COURSE_ID", "")
        if cid:
            api_info = _fetch_course_from_canvas(cid)
            name = api_info.get("name", "")
            term = api_info.get("term", "")
            if name:
                return name, term or "No-Term", name
    except Exception:
        pass

    import logging as _logging
    _logging.getLogger(__name__).warning(
        "Could not determine course name/code — all fallbacks failed. "
        "Ensure course-config.json has 'course_code' and 'term' fields, "
        "or that CANVAS_COURSE_ID in .env matches your course."
    )
    return "UNKNOWN", "No-Term", ""


def _fetch_course_from_canvas(course_id: str) -> dict:
    """Fetch course name and term from Canvas API.

    Returns dict with 'name', 'course_code', 'term' keys (empty strings on failure).
    """
    result = {"name": "", "course_code": "", "term": ""}
    try:
        from dotenv import load_dotenv
        load_dotenv(PLUGIN_ROOT / ".env")
        load_dotenv(PLUGIN_ROOT / ".env.local", override=True)
        import requests
        token = os.getenv("CANVAS_TOKEN")
        domain = os.getenv("CANVAS_DOMAIN")
        if not token or not domain:
            return result
        r = requests.get(
            f"https://{domain}/api/v1/courses/{course_id}",
            headers={"Authorization": f"Bearer {token}"},
            params={"include[]": "term"},
            timeout=10,
        )
        if r.status_code == 200:
            data = r.json()
            result["name"] = data.get("name", "")
            result["course_code"] = data.get("course_code", "")
            term_data = data.get("term", {})
            if isinstance(term_data, dict):
                result["term"] = term_data.get("name", "")
        return result
    except Exception:
        return result


def _sanitize_folder_name(name: str) -> str:
    """Sanitize a string for use as a filesystem folder/file name."""
    import re as _re
    # Replace characters not safe for filesystems
    safe = _re.sub(r'[<>:"/\\|?*]', '-', name)
    # Collapse multiple hyphens/underscores/spaces
    safe = _re.sub(r'[\s]+', '-', safe)
    # Strip leading/trailing dots and whitespace
    return safe.strip('. ')


def _find_existing_course_folder(course_name: str, term: str) -> Path | None:
    """Find an existing reports subfolder that matches the course name + term.

    Normalizes hyphens/underscores/spaces for comparison.
    Returns the existing folder Path, or None if no match.
    """
    import re as _re
    if not REPORTS_DIR.exists():
        return None
    target = _re.sub(r"[-_\s]", "", f"{course_name}{term}").upper()
    for folder in REPORTS_DIR.iterdir():
        if folder.is_dir():
            normalized = _re.sub(r"[-_\s]", "", folder.name).upper()
            if normalized == target:
                return folder
    return None


def _build_report_path(ext="html", data: dict = None):
    """Build a timestamped, course-scoped report path under reports/.

    Structure: reports/{CourseName_Term}/{CourseName_YYYY-MM-DD_HH-MM_AI-Audit}.{ext}
    Uses the exact Canvas course name — no regex parsing or normalization.
    Reuses existing matching folder if found.
    """
    course_name, term, _ = _get_course_info(data)
    safe_name = _sanitize_folder_name(course_name)
    safe_term = _sanitize_folder_name(term)
    folder_name = f"{safe_name}_{safe_term}"
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"{safe_name}_{timestamp}_AI-Audit.{ext}"

    # Try to reuse an existing folder that matches
    existing = _find_existing_course_folder(course_name, term)
    report_dir = existing if existing else (REPORTS_DIR / folder_name)
    report_dir.mkdir(parents=True, exist_ok=True)
    return report_dir / filename


def _escape(text: str) -> str:
    """HTML-escape text."""
    return (str(text)
            .replace('&', '&amp;')
            .replace('<', '&lt;')
            .replace('>', '&gt;')
            .replace('"', '&quot;'))


def _severity_color(severity: str) -> str:
    """Return color for severity level."""
    s = severity.lower()
    if s in ('met', 'pass', 'info'):
        return '#1e7e34'
    elif s in ('partially met', 'warn', 'warning'):
        return '#b5540a'
    elif s in ('not met', 'fail', 'critical'):
        return '#c62828'
    elif s in ('not auditable',):
        return '#6a7883'
    return '#6a7883'


def _severity_bg(severity: str) -> str:
    s = severity.lower()
    if s in ('met', 'pass', 'info'):
        return '#e6f4ea'
    elif s in ('partially met', 'warn', 'warning'):
        return '#fef3e5'
    elif s in ('not met', 'fail', 'critical'):
        return '#fde8e8'
    elif s in ('not auditable',):
        return '#eef0f2'
    return '#f5f5f5'


def _severity_icon(severity: str) -> str:
    s = severity.lower()
    if s in ('met', 'pass', 'info'):
        return '✓'
    elif s in ('partially met', 'warn', 'warning'):
        return '⚠'
    elif s in ('not met', 'fail', 'critical'):
        return '✗'
    elif s in ('not auditable',):
        return '○'
    return '–'


def _render_criteria_results(criteria_results: list) -> str:
    """Render per-criterion results as expandable sub-rows within a standard card."""
    if not criteria_results:
        return ''
    met_count = sum(1 for cr in criteria_results if cr.get('status', '').lower() in ('met', 'pass', 'yes'))
    total = len(criteria_results)
    rows = []
    for cr in criteria_results:
        cid = _escape(cr.get('criterion_id', ''))
        status = cr.get('status', 'Unknown')
        ct = cr.get('check_type', 'ai')
        evidence = _escape(cr.get('evidence', ''))
        tier = cr.get('reviewer_tier', '')
        tier_badge = ''
        if cid.startswith('B-') or tier == 'id_assistant':
            tier_badge = '<span style="background:#e0f4fa;color:#0081b3;font-size:10px;padding:1px 4px;border-radius:3px;margin-left:4px">B</span>'
        elif cid.startswith('C-') or tier == 'id':
            tier_badge = '<span style="background:#f3ebe7;color:#AF674B;font-size:10px;padding:1px 4px;border-radius:3px;margin-left:4px">C</span>'
        graph_tag = ' 📊' if cr.get('graph_verified') else ''
        type_label = {'deterministic': '⚙ Auto', 'ai': '🤖 AI', 'hybrid': '⚙+🤖'}.get(ct, ct)
        rows.append(
            f'<tr style="font-size:12px">'
            f'<td style="padding:4px 8px;color:#888;white-space:nowrap">{cid}{tier_badge}</td>'
            f'<td style="padding:4px 8px"><span style="color:{_severity_color(status)};font-weight:600">'
            f'{_severity_icon(status)} {_escape(status)}</span></td>'
            f'<td style="padding:4px 8px;color:#888">{type_label}{graph_tag}</td>'
            f'<td style="padding:4px 8px">{evidence[:150]}{"..." if len(evidence) > 150 else ""}</td>'
            f'</tr>'
        )
    has_issues = any(cr.get('status') in ('Not Met', 'Partially Met', 'not_met', 'not met', 'No') for cr in criteria_results)
    open_attr = ' open' if has_issues else ''
    summary_color = '#1e7e34' if met_count == total else ('#b5540a' if met_count > total / 2 else '#c62828')
    return (
        f'<details style="margin-top:8px;border-top:1px solid #eee;padding-top:6px"{open_attr}>'
        f'<summary style="cursor:pointer;font-size:12px;color:#888">'
        f'Criteria: <span style="color:{summary_color};font-weight:600">{met_count}/{total} met</span></summary>'
        f'<table style="width:100%;margin-top:6px;border-collapse:collapse">'
        f'<thead><tr style="font-size:11px;color:#aaa"><th style="text-align:left;padding:2px 8px">ID</th>'
        f'<th style="text-align:left;padding:2px 8px">Status</th>'
        f'<th style="text-align:left;padding:2px 8px">Mode</th>'
        f'<th style="text-align:left;padding:2px 8px">Evidence</th></tr></thead>'
        f'<tbody>{"".join(rows)}</tbody></table></details>'
    )


def generate_demo_data() -> dict:
    """Generate realistic sample audit data for demo/testing."""
    return {
        "course": {
            "name": "BIO 201 — Human Physiology",
            "id": "218764",
            "institution": "Arizona State University",
            "domain": "canvas.asu.edu"
        },
        "audit_date": datetime.now().isoformat(),
        "auditor": "ID Workbench v1.0.2",
        "sections": {
            "design_standards": {
                "title": "ASU Course Design Standards",
                "subtitle": "25 standards evaluated against measurable criteria",
                "summary": {"Met": 18, "Partially Met": 5, "Not Met": 2},
                "items": [
                    {"id": "01", "name": "Course-Level Alignment", "category": "Course Structure and Organization", "status": "Met", "evidence": "All 5 CLOs use measurable action verbs (analyze, evaluate, apply, compare, design) and align with BIO program objectives.", "recommendation": None, "confidence": "High", "coverage": "Course-wide", "scope": "Course-wide", "evidence_source": "Canvas", "canvas_link": "https://canvas.asu.edu/courses/218764/pages/syllabus", "reviewer_tier": "id"},
                    {"id": "02", "name": "Module-Level Alignment", "category": "Course Structure and Organization", "status": "Met", "evidence": "Each module contains 3-5 objectives mapped to CLOs via the alignment matrix on overview pages.", "recommendation": None, "confidence": "High", "coverage": "13/13 modules", "scope": "Module-level", "evidence_source": "Canvas", "canvas_link": "https://canvas.asu.edu/courses/218764/pages/m1-overview", "reviewer_tier": "id"},
                    {"id": "03", "name": "Alignment Made Clear", "category": "Course Structure and Organization", "status": "Partially Met", "evidence": "Overview pages list objectives but Modules 4-5 don't explicitly connect assessments to specific CLOs.", "recommendation": "Add CLO tags to assessment descriptions in Modules 4 and 5.", "confidence": "Medium", "coverage": "11/13 modules", "scope": "Module-level", "evidence_source": "Canvas", "reviewer_tier": "id_assistant"},
                    {"id": "04", "name": "Consistent Layout", "category": "Course Structure and Organization", "status": "Met", "evidence": "All 13 modules follow the 7-page structure: Overview → Prepare → Lesson → Practice → Knowledge Check → Artifact → Conclusion.", "recommendation": None, "confidence": "High", "coverage": "13/13 modules", "scope": "Module-level", "evidence_source": "Canvas"},
                    {"id": "05", "name": "Engaging Introductions", "category": "Course Structure and Organization", "status": "Met", "evidence": "Module overviews include real-world scenarios, learning roadmaps, and professional connection callouts.", "recommendation": None, "confidence": "High", "coverage": "13/13 modules", "scope": "Module-level", "evidence_source": "Canvas"},
                    {"id": "06", "name": "Balanced Workload", "category": "Course Structure and Organization", "status": "Met", "evidence": "Module workload ranges from 8.5-11.2 hours across all 13 modules, within the 7.5-week session guidelines.", "recommendation": None, "confidence": "High", "coverage": "13/13 modules", "scope": "Course-wide", "evidence_source": "Canvas"},
                    {"id": "07", "name": "Clear Expectations", "category": "Learner Support", "status": "Met", "evidence": "Syllabus includes due dates, late policy, grading scale, and participation expectations. Each assessment has a rubric.", "recommendation": None, "confidence": "High", "coverage": "Course-wide", "scope": "Course-wide", "evidence_source": "Canvas"},
                    {"id": "08", "name": "Active Learning", "category": "Learner Engagement", "status": "Met", "evidence": "Guided practice activities in every module include interactive H5P elements (drag-drop, sequencing, fill-in-the-blank).", "recommendation": None, "confidence": "High", "coverage": "13/13 modules", "scope": "Module-level", "evidence_source": "Canvas"},
                    {"id": "09", "name": "Multiple Modalities", "category": "Learner Engagement", "status": "Partially Met", "evidence": "Modules 1-10 have video lectures + text + interactive activities. Modules 11-13 lack video content.", "recommendation": "Add video lectures or narrated walkthroughs to Modules 11-13.", "confidence": "High", "coverage": "10/13 modules", "scope": "Module-level", "evidence_source": "Canvas"},
                    {"id": "10", "name": "Learner Interaction", "category": "Learner Engagement", "status": "Met", "evidence": "Weekly discussion forums with peer response requirements. Group artifact in Module 7.", "recommendation": None, "confidence": "High", "coverage": "13/13 modules", "scope": "Module-level", "evidence_source": "Canvas"},
                    {"id": "11", "name": "Assessment Variety", "category": "Assessment", "status": "Met", "evidence": "Course uses knowledge checks (auto-graded), guided practice, discussions, creative artifacts, and a cumulative final.", "recommendation": None, "confidence": "High", "coverage": "Course-wide", "scope": "Course-wide", "evidence_source": "Canvas"},
                    {"id": "12", "name": "Formative Assessment", "category": "Assessment", "status": "Met", "evidence": "Knowledge checks allow 3 attempts with feedback. Guided practice is ungraded with immediate feedback.", "recommendation": None, "confidence": "High", "coverage": "13/13 modules", "scope": "Module-level", "evidence_source": "Canvas"},
                    {"id": "13", "name": "Rubric Quality", "category": "Assessment", "status": "Partially Met", "evidence": "Rubrics present for all artifacts and discussions. Module 6 rubric uses vague criteria ('good effort').", "recommendation": "Revise Module 6 artifact rubric with specific, measurable criteria.", "confidence": "Medium", "coverage": "12/13 modules", "scope": "Assessment-level", "evidence_source": "Canvas"},
                    {"id": "14", "name": "Feedback Opportunities", "category": "Assessment", "status": "Met", "evidence": "Auto-feedback on quizzes, rubric-based feedback on artifacts, peer feedback in discussions.", "recommendation": None, "confidence": "High", "coverage": "13/13 modules", "scope": "Module-level", "evidence_source": "Canvas"},
                    {"id": "15", "name": "Academic Integrity", "category": "Assessment", "status": "Met", "evidence": "Academic integrity agreement in Module 0. Turnitin enabled on artifact submissions. AI policy clearly stated.", "recommendation": None, "confidence": "High", "coverage": "Course-wide", "scope": "Course-wide", "evidence_source": "Canvas"},
                    {"id": "16", "name": "Accessible Content", "category": "Accessibility & Usability", "status": "Partially Met", "evidence": "Most images have alt text. 3 decorative images incorrectly have descriptive alt text. 2 complex diagrams lack long descriptions.", "recommendation": "Add alt=\"\" to decorative images. Add long descriptions for anatomy diagrams in Modules 2 and 8.", "confidence": "High", "coverage": "11/13 modules", "scope": "Module-level", "evidence_source": "Canvas"},
                    {"id": "17", "name": "Navigable Structure", "category": "Accessibility & Usability", "status": "Met", "evidence": "Consistent heading hierarchy (H2→H3→H4). No skipped levels detected.", "recommendation": None, "confidence": "High", "coverage": "13/13 modules", "scope": "Module-level", "evidence_source": "Canvas"},
                    {"id": "18", "name": "Media Accessibility", "category": "Accessibility & Usability", "status": "Not Met", "evidence": "Video captions present for Modules 1-7 only. Modules 8-10 videos have auto-generated captions that haven't been reviewed.", "recommendation": "Review and correct auto-generated captions for Modules 8-10 videos.", "confidence": "High", "coverage": "7/13 modules", "scope": "Module-level", "evidence_source": "Canvas"},
                    {"id": "19", "name": "Color Independence", "category": "Accessibility & Usability", "status": "Met", "evidence": "No information conveyed by color alone. Charts use patterns + color. All links underlined.", "recommendation": None, "confidence": "High", "coverage": "13/13 modules", "scope": "Module-level", "evidence_source": "Canvas"},
                    {"id": "20", "name": "Inclusive Design", "category": "Accessibility & Usability", "status": "Met", "evidence": "Content uses gender-neutral language. Case studies represent diverse populations and clinical settings.", "recommendation": None, "confidence": "Medium", "coverage": "13/13 modules", "scope": "Module-level", "evidence_source": "Canvas"},
                    {"id": "21", "name": "Current Content", "category": "Content Quality", "status": "Partially Met", "evidence": "Most references are 2022-2025. Module 3 cites a 2018 study as 'recent research.'", "recommendation": "Update Module 3 reference or remove 'recent' qualifier.", "confidence": "Medium", "coverage": "12/13 modules", "scope": "Module-level", "evidence_source": "Canvas"},
                    {"id": "22", "name": "Credible Sources", "category": "Content Quality", "status": "Met", "evidence": "All external sources are peer-reviewed journals, established textbooks, or professional organization publications.", "recommendation": None, "confidence": "Medium", "coverage": "Course-wide", "scope": "Course-wide", "evidence_source": "Mixed"},
                    {"id": "23", "name": "Appropriate Depth", "category": "Content Quality", "status": "Met", "evidence": "Content progresses from foundational (Modules 1-4) to applied (5-9) to integrative (10-13), appropriate for a 200-level course.", "recommendation": None, "confidence": "Medium", "coverage": "Course-wide", "scope": "Course-wide", "evidence_source": "Canvas"},
                    {"id": "24", "name": "Technology Integration", "category": "Technology", "status": "Not Met", "evidence": "H5P activities work in Chrome and Firefox but Module 9 interactive fails on Safari. External link to PhET simulation returns 404.", "recommendation": "Fix Safari compatibility for Module 9 H5P. Replace broken PhET link.", "confidence": "High", "coverage": "1/13 modules", "scope": "Module-level", "evidence_source": "External"},
                    {"id": "25", "name": "Privacy & FERPA", "category": "Technology", "status": "Met", "evidence": "All tools are ASU-approved. No student data shared with third parties. Proctoring tool has approved DPA.", "recommendation": None, "confidence": "Low", "coverage": "Course-wide", "scope": "Course-wide", "evidence_source": "External"},
                ]
            },
            "qa_categories": {
                "title": "QA Categories",
                "subtitle": "19 structural and formatting checks",
                "summary": {"Pass": 14, "Warn": 3, "Fail": 2},
                "items": [
                    {"id": "Q01", "name": "Module structure consistency", "status": "Pass", "detail": "All 13 modules follow identical 7-page structure."},
                    {"id": "Q02", "name": "Page title formatting", "status": "Pass", "detail": "All titles follow 'Module X: Title — Page Type' convention."},
                    {"id": "Q03", "name": "Learning objectives presence", "status": "Warn", "detail": "Modules 1-11 have objectives on overview. Modules 12-13 list objectives in lesson pages instead."},
                    {"id": "Q04", "name": "Assessment-objective alignment", "status": "Pass", "detail": "All assessments reference at least one module objective."},
                    {"id": "Q05", "name": "Quiz configuration", "status": "Pass", "detail": "All quizzes: 3 attempts, shuffle answers, 30-min time limit."},
                    {"id": "Q06", "name": "Heading hierarchy", "status": "Pass", "detail": "No heading skips detected across 91 pages."},
                    {"id": "Q07", "name": "Image alt text presence", "status": "Warn", "detail": "96% coverage. 3 images missing alt text in Modules 2, 8."},
                    {"id": "Q08", "name": "Link accessibility", "status": "Pass", "detail": "All external links use target='_blank' with noopener and descriptive text."},
                    {"id": "Q09", "name": "Video captions/transcripts", "status": "Fail", "detail": "Modules 8-10 have unreviewed auto-captions. No transcripts for Modules 11-13 audio content."},
                    {"id": "Q10", "name": "File accessibility", "status": "Pass", "detail": "All PDFs are tagged. No scanned-image PDFs detected."},
                    {"id": "Q11", "name": "Color contrast", "status": "Pass", "detail": "All inline text colors meet WCAG 2.1 AA 4.5:1 ratio."},
                    {"id": "Q12", "name": "Table structure", "status": "Pass", "detail": "All 18 tables have proper thead, th elements, and scope attributes."},
                    {"id": "Q13", "name": "Font consistency", "status": "Pass", "detail": "All pages use Roboto via Canvas theme. No inline font overrides."},
                    {"id": "Q14", "name": "Navigation consistency", "status": "Pass", "detail": "Standard Canvas sidebar. Module navigation tabs cleaned up."},
                    {"id": "Q15", "name": "Module completeness", "status": "Pass", "detail": "All modules contain all 7 required page types."},
                    {"id": "Q16", "name": "Grading transparency", "status": "Pass", "detail": "Assignment groups with weights. All items have point values. Grading scale in syllabus."},
                    {"id": "Q17", "name": "External link validation", "status": "Fail", "detail": "1 broken link: PhET simulation in Module 9 returns 404."},
                    {"id": "Q18", "name": "Content freshness", "status": "Warn", "detail": "Module 3 references 2018 study as 'recent.' Module 7 mentions 'last year' without specific date."},
                    {"id": "Q19", "name": "Mobile-friendly layout", "status": "Pass", "detail": "All pages render correctly at 375px viewport. No horizontal scroll detected."},
                ]
            },
            "accessibility": {
                "title": "WCAG 2.1 AA Accessibility",
                "subtitle": "Automated accessibility checks across all pages",
                "summary": {"Critical": 2, "Warning": 5, "Info": 3},
                "items": [
                    {"severity": "Critical", "page": "m8-lesson-introduction", "issue": "Image missing alt text", "element": "<img src='anatomy-diagram-8.png'>", "fix": "Add descriptive alt text: 'Diagram of the musculoskeletal system showing major muscle groups and bone attachment points'", "canvas_link": "https://canvas.asu.edu/courses/218764/pages/m8-lesson-introduction"},
                    {"severity": "Critical", "page": "m2-lesson-introduction", "issue": "Image missing alt text", "element": "<img src='cardiac-cycle.png'>", "fix": "Add descriptive alt text describing the cardiac cycle diagram phases", "canvas_link": "https://canvas.asu.edu/courses/218764/pages/m2-lesson-introduction"},
                    {"severity": "Warning", "page": "m6-guided-practice", "issue": "Generic link text", "element": "<a href='...'>click here</a>", "fix": "Replace with descriptive text: 'Open the endocrine system simulation'"},
                    {"severity": "Warning", "page": "m3-resources", "issue": "Generic link text", "element": "<a href='...'>read more</a>", "fix": "Replace with: 'Read the full research article on neural pathways'"},
                    {"severity": "Warning", "page": "m9-lesson-introduction", "issue": "Low contrast inline color", "element": "<span style='color:#999'>", "fix": "Change to #595959 or darker for 4.5:1 contrast ratio"},
                    {"severity": "Warning", "page": "m12-overview", "issue": "Heading hierarchy skip", "element": "h2 → h4", "fix": "Change h4 to h3 to maintain proper hierarchy"},
                    {"severity": "Warning", "page": "m5-conclusion", "issue": "Heading hierarchy skip", "element": "h2 → h4", "fix": "Change h4 to h3"},
                    {"severity": "Info", "page": "m1-overview", "issue": "Decorative image has alt text", "element": "<img alt='decorative banner'>", "fix": "Change to alt='' for decorative images"},
                    {"severity": "Info", "page": "m7-overview", "issue": "Decorative image has alt text", "element": "<img alt='module header image'>", "fix": "Change to alt='' for decorative images"},
                    {"severity": "Info", "page": "m10-resources", "issue": "Long alt text (>125 chars)", "element": "<img alt='Detailed diagram showing...'>", "fix": "Move to longdesc or aria-describedby for complex images"},
                ]
            },
            "readiness": {
                "title": "Course Readiness Check",
                "subtitle": "Pre-launch operational checklist — 9 categories",
                "overall": "NOT READY",
                "categories": [
                    {"name": "Course Information", "status": "Pass", "checks": [
                        {"item": "Syllabus published", "status": "Pass"},
                        {"item": "Course description set", "status": "Pass"},
                        {"item": "Instructor info visible", "status": "Pass"},
                    ]},
                    {"name": "Navigation", "status": "Pass", "checks": [
                        {"item": "Home page configured", "status": "Pass"},
                        {"item": "Navigation tabs cleaned up", "status": "Pass"},
                        {"item": "Modules ordered", "status": "Pass"},
                    ]},
                    {"name": "Content Availability", "status": "Pass", "checks": [
                        {"item": "All modules have content", "status": "Pass"},
                        {"item": "No empty pages", "status": "Pass"},
                        {"item": "No placeholder text", "status": "Pass"},
                    ]},
                    {"name": "Assessment Setup", "status": "Pass", "checks": [
                        {"item": "Quiz attempts configured", "status": "Pass"},
                        {"item": "Quiz time limits set", "status": "Pass"},
                        {"item": "All assignments have due dates", "status": "Pass"},
                        {"item": "Rubrics attached", "status": "Pass"},
                    ]},
                    {"name": "Grading", "status": "Pass", "checks": [
                        {"item": "Assignment groups with weights", "status": "Pass"},
                        {"item": "Grading scheme set", "status": "Pass"},
                        {"item": "Late policy configured", "status": "Pass"},
                    ]},
                    {"name": "Dates & Availability", "status": "Fail", "checks": [
                        {"item": "All due dates set", "status": "Fail", "note": "Module 13 artifact missing due date"},
                        {"item": "No past dates", "status": "Pass"},
                        {"item": "Module availability dates", "status": "Pass"},
                    ]},
                    {"name": "Communication", "status": "Pass", "checks": [
                        {"item": "Welcome announcement drafted", "status": "Pass"},
                        {"item": "Discussion boards ready", "status": "Pass"},
                        {"item": "Instructor contact info", "status": "Pass"},
                    ]},
                    {"name": "Publishing", "status": "Fail", "checks": [
                        {"item": "All modules published", "status": "Fail", "note": "Module 13 still in draft state"},
                        {"item": "No draft content published", "status": "Pass"},
                    ]},
                    {"name": "Technical", "status": "Fail", "checks": [
                        {"item": "No broken links", "status": "Fail", "note": "1 broken link in Module 9"},
                        {"item": "All media playable", "status": "Pass"},
                        {"item": "External tools configured", "status": "Pass"},
                    ]},
                ]
            }
        },
        "clo_alignment": {
            "title": "CLO Alignment Matrix",
            "clos": [
                {"id": "CLO-1", "text": "Analyze the structure and function of major organ systems", "modules": [1,2,3,4,5], "assessments": 14},
                {"id": "CLO-2", "text": "Evaluate homeostatic mechanisms in health and disease", "modules": [3,4,5,6,7], "assessments": 12},
                {"id": "CLO-3", "text": "Apply physiological principles to clinical case studies", "modules": [6,7,8,9,10], "assessments": 10},
                {"id": "CLO-4", "text": "Compare regulatory systems across organ systems", "modules": [8,9,10,11], "assessments": 8},
                {"id": "CLO-5", "text": "Design evidence-based explanations of physiological phenomena", "modules": [10,11,12,13], "assessments": 6},
            ]
        },
        "external_links": [
            {"page": "m3-resources", "text": "NIH Research Article", "url": "https://www.nih.gov/research/physiology-2024", "domain": "nih.gov", "status": "Not Reviewed", "notes": "External — not in audit scope"},
            {"page": "m5-lesson-introduction", "text": "Khan Academy Video", "url": "https://www.khanacademy.org/science/biology/circulatory-system", "domain": "khanacademy.org", "status": "Not Reviewed", "notes": "External — not in audit scope"},
            {"page": "m9-lesson-introduction", "text": "PhET Simulation", "url": "https://phet.colorado.edu/en/simulation/legacy/nerve-conduction", "domain": "phet.colorado.edu", "status": "Broken", "notes": "Returns 404 — link needs replacement"},
            {"page": "m7-guided-practice", "text": "PubMed Central", "url": "https://www.ncbi.nlm.nih.gov/pmc/articles/PMC12345/", "domain": "ncbi.nlm.nih.gov", "status": "Not Reviewed", "notes": "External — not in audit scope"},
            {"page": "m1-overview", "text": "ASU Biology Department", "url": "https://sols.asu.edu/degree-programs/biology", "domain": "asu.edu", "status": "Reviewed", "notes": "Trusted domain — verified accessible"},
        ]
    }


def _normalize_audit_data(data: dict) -> dict:
    """Normalize audit data to expected schema.

    Fixes common mismatches:
    1. Sections at top level instead of nested under 'sections'
    2. Lowercase summary keys (met -> Met, pass -> Pass)
    3. Missing 'items' arrays
    """
    SECTION_KEYS = ("design_standards", "qa_categories", "accessibility", "readiness")

    # Fix 1: Unwrapped sections — top-level keys need wrapping
    if not data.get("sections"):
        wrapped = {}
        for key in SECTION_KEYS:
            if key in data:
                wrapped[key] = data.pop(key)
        if wrapped:
            data["sections"] = wrapped

    if not data.get("sections"):
        data["sections"] = {}

    sections = data["sections"]

    # Fix 2: Normalize summary key casing
    CASE_MAP = {
        "met": "Met", "partially met": "Partially Met", "not met": "Not Met",
        "not auditable": "Not Auditable",
        "pass": "Pass", "warn": "Warn", "fail": "Fail",
        "critical": "Critical", "warning": "Warning", "info": "Info",
    }
    for key in SECTION_KEYS:
        section = sections.get(key, {})
        if "summary" in section:
            normalized = {}
            for k, v in section["summary"].items():
                normalized[CASE_MAP.get(k.lower(), k)] = v
            section["summary"] = normalized

    # Fix 3: Ensure 'items' arrays exist
    for key in ("design_standards", "qa_categories", "accessibility"):
        section = sections.get(key, {})
        if "items" not in section:
            section["items"] = []
        sections[key] = section

    # Readiness uses 'categories' not 'items'
    readiness = sections.get("readiness", {})
    if "categories" not in readiness:
        readiness["categories"] = []
    sections["readiness"] = readiness

    return data


def _build_remediation_html(sections: dict, overall_score: float) -> str:
    """Build the Remediation Roadmap section HTML from audit findings.

    Auto-ranks all Not Met / Partially Met / Fail / Warn items by estimated score
    impact and emits a collapsible section with progress bars and finding cards.
    """
    ds_items  = sections.get('design_standards', {}).get('items', [])
    qa_items  = sections.get('qa_categories', {}).get('items', [])
    a11y_items = sections.get('accessibility', {}).get('items', [])
    r_cats    = sections.get('readiness', {}).get('categories', [])

    ds_total  = max(len(ds_items), 1)
    qa_total  = max(len(qa_items), 1)
    r_total   = max(len(r_cats), 1)

    # Weight contribution per item (mirrors overall scoring: DS 40%, QA 30%, A11y 15%, Readiness 15%)
    ds_per    = 40.0 / ds_total
    qa_per    = 30.0 / qa_total
    r_per     = 15.0 / r_total

    remediations = []

    # ── Design Standards: Not Met ───────────────────────────────────────────
    for item in ds_items:
        if item.get('status') == 'Not Met':
            rec = item.get('recommendation') or 'Address this standard to meet ASU requirements.'
            remediations.append({
                'priority': 'critical',
                'label': '🔴 Critical',
                'badge_bg': '#fde8e8', 'badge_fg': '#c62828',
                'border': '#c62828',
                'source': f"DS-{item.get('id', '')}",
                'title': item.get('name', ''),
                'detail': rec,
                'impact': round(ds_per, 1),
            })
        elif item.get('status') == 'Partially Met':
            rec = item.get('recommendation') or 'Complete remaining criteria to fully meet this standard.'
            remediations.append({
                'priority': 'medium',
                'label': '🟡 Medium',
                'badge_bg': '#fffde7', 'badge_fg': '#f9a825',
                'border': '#f9a825',
                'source': f"DS-{item.get('id', '')}",
                'title': item.get('name', ''),
                'detail': rec,
                'impact': round(ds_per * 0.5, 1),
            })

    # ── QA Categories: Fail then Warn ───────────────────────────────────────
    for item in qa_items:
        st = item.get('status', '')
        if st == 'Fail':
            remediations.append({
                'priority': 'critical',
                'label': '🔴 Critical',
                'badge_bg': '#fde8e8', 'badge_fg': '#c62828',
                'border': '#c62828',
                'source': 'QA',
                'title': item.get('name', ''),
                'detail': item.get('detail') or item.get('description') or 'Fix this QA failure before launch.',
                'impact': round(qa_per, 1),
            })
        elif st == 'Warn':
            remediations.append({
                'priority': 'high',
                'label': '🟠 High',
                'badge_bg': '#fff3e0', 'badge_fg': '#e65100',
                'border': '#e65100',
                'source': 'QA',
                'title': item.get('name', ''),
                'detail': item.get('detail') or item.get('description') or 'Address this warning before launch.',
                'impact': round(qa_per * 0.5, 1),
            })

    # ── Accessibility ────────────────────────────────────────────────────────
    for item in a11y_items:
        sev = item.get('severity', '')
        if sev == 'Critical':
            remediations.append({
                'priority': 'critical',
                'label': '🔴 Critical',
                'badge_bg': '#fde8e8', 'badge_fg': '#c62828',
                'border': '#c62828',
                'source': 'WCAG',
                'title': f"[Accessibility] {item.get('issue', '')}",
                'detail': item.get('fix') or item.get('issue', ''),
                'impact': 3.0,
            })
        elif sev == 'Warning':
            remediations.append({
                'priority': 'high',
                'label': '🟠 High',
                'badge_bg': '#fff3e0', 'badge_fg': '#e65100',
                'border': '#e65100',
                'source': 'WCAG',
                'title': f"[Accessibility] {item.get('issue', '')}",
                'detail': item.get('fix') or item.get('issue', ''),
                'impact': 1.5,
            })

    # ── Readiness: Fail then Partial ─────────────────────────────────────────
    for cat in r_cats:
        st = cat.get('status', '')
        if st == 'Fail':
            checks_txt = '; '.join(
                c.get('label', '') for c in cat.get('checks', []) if c.get('status') != 'Pass'
            ) or 'Complete all checks in this category.'
            remediations.append({
                'priority': 'high',
                'label': '🟠 High',
                'badge_bg': '#fff3e0', 'badge_fg': '#e65100',
                'border': '#e65100',
                'source': f"Readiness — {cat.get('name', '')}",
                'title': cat.get('name', ''),
                'detail': checks_txt,
                'impact': round(r_per, 1),
            })
        elif st == 'Partial':
            remediations.append({
                'priority': 'medium',
                'label': '🟡 Medium',
                'badge_bg': '#fffde7', 'badge_fg': '#f9a825',
                'border': '#f9a825',
                'source': f"Readiness — {cat.get('name', '')}",
                'title': cat.get('name', ''),
                'detail': 'Complete remaining checks in this category.',
                'impact': round(r_per * 0.5, 1),
            })

    if not remediations:
        return '''
    <div class="report-section" id="section-remediation">
      <div class="section-header" onclick="this.parentElement.classList.toggle('collapsed')">
        <div>
          <h2>Remediation Roadmap</h2>
          <div class="section-subtitle">No issues found — course meets all evaluated standards</div>
        </div>
        <span class="section-toggle">▾</span>
      </div>
      <div class="section-body">
        <div style="padding:32px;text-align:center;color:#1e7e34;font-size:15px">
          ✅ All audited standards are Met or Pass. No remediation required.
        </div>
      </div>
    </div>'''

    # Sort: critical first, then by impact desc
    priority_order = {'critical': 0, 'high': 1, 'medium': 2}
    remediations.sort(key=lambda x: (priority_order.get(x['priority'], 9), -x['impact']))

    projected = min(round(overall_score + sum(r['impact'] for r in remediations)), 100)
    current_pct  = round(overall_score)
    current_w    = max(min(current_pct, 100), 1)
    projected_w  = max(min(projected, 100), 1)
    current_col  = '#c62828' if current_pct < 70 else '#b5540a' if current_pct < 80 else '#1e7e34'
    projected_col = '#1e7e34' if projected >= 80 else '#b5540a'

    cards_html = []
    for idx, r in enumerate(remediations, 1):
        cards_html.append(f'''
        <div class="finding-card" data-status="not-met" style="border-left:4px solid {r["border"]}">
          <div style="display:flex;align-items:flex-start;justify-content:space-between;gap:12px">
            <div style="flex:1">
              <div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">
                <span style="background:{r["badge_bg"]};color:{r["badge_fg"]};font-weight:700;padding:2px 10px;border-radius:12px;font-size:12px">{r["label"]}</span>
                <span style="background:#f5f5f5;color:#555;padding:2px 8px;border-radius:12px;font-size:11px">#{idx}</span>
                <span style="font-size:12px;color:#6a7883">{_escape(r["source"])}</span>
              </div>
              <div style="font-weight:600;font-size:15px;margin-bottom:4px">{_escape(r["title"])}</div>
              <div style="font-size:13px;color:var(--canvas-text-secondary)">{_escape(r["detail"])}</div>
            </div>
            <div style="text-align:right;white-space:nowrap">
              <div style="font-size:20px;font-weight:700;color:#1e7e34">+{r["impact"]} pts</div>
            </div>
          </div>
        </div>''')

    quick_wins = [r for r in remediations if r['priority'] == 'critical']
    qw_pts = round(sum(r['impact'] for r in quick_wins), 1)
    qw_note = (
        f'<strong>Quick wins:</strong> Fixing the {len(quick_wins)} critical item{"s" if len(quick_wins) != 1 else ""} '
        f'adds ~{qw_pts} pts to your score.'
    ) if quick_wins else ''

    return f'''
    <!-- Remediation Roadmap -->
    <div class="report-section" id="section-remediation">
      <div class="section-header" onclick="this.parentElement.classList.toggle('collapsed')">
        <div>
          <h2>Remediation Roadmap</h2>
          <div class="section-subtitle">{len(remediations)} action{"s" if len(remediations) != 1 else ""} identified — ranked by priority and score impact</div>
        </div>
        <span class="section-toggle">▾</span>
      </div>
      <div class="section-body">
        <div style="padding:20px 24px 0">
          <div style="display:flex;align-items:center;gap:16px;margin-bottom:8px">
            <span style="font-size:13px;color:var(--canvas-text-secondary);min-width:120px">Current score</span>
            <div style="flex:1;background:#eee;border-radius:4px;height:10px">
              <div style="width:{current_w}%;background:{current_col};height:10px;border-radius:4px"></div>
            </div>
            <span style="font-weight:700;color:{current_col};min-width:48px">{current_pct}%</span>
          </div>
          <div style="display:flex;align-items:center;gap:16px;margin-bottom:20px">
            <span style="font-size:13px;color:var(--canvas-text-secondary);min-width:120px">After remediation</span>
            <div style="flex:1;background:#eee;border-radius:4px;height:10px">
              <div style="width:{projected_w}%;background:{projected_col};height:10px;border-radius:4px"></div>
            </div>
            <span style="font-weight:700;color:{projected_col};min-width:48px">{projected}%</span>
          </div>
          {'<p style="font-size:13px;color:var(--canvas-text-secondary);margin-bottom:4px">' + qw_note + '</p>' if qw_note else ''}
        </div>
        {''.join(cards_html)}
        <div style="padding:16px 24px;background:#e8f5e9;border-top:2px solid #1e7e34;display:flex;align-items:center;justify-content:space-between">
          <div>
            <div style="font-weight:700;font-size:15px;color:#1e7e34">Projected score after all remediations</div>
            <div style="font-size:13px;color:#555">{len(remediations)} items · sorted by priority and estimated impact</div>
          </div>
          <div style="font-size:28px;font-weight:700;color:#1e7e34">{projected}%</div>
        </div>
      </div>
    </div>'''


def generate_report(data: dict) -> str:
    """Generate the full HTML audit report."""
    data = _normalize_audit_data(data)
    course = data.get('course', {})
    audit_date = data.get('audit_date', datetime.now().isoformat())
    auditor = _resolve_auditor(data)
    sections = data.get('sections', {})
    clo_data = data.get('clo_alignment', {})

    try:
        date_display = datetime.fromisoformat(audit_date).strftime('%B %d, %Y at %I:%M %p')
    except (ValueError, TypeError):
        date_display = audit_date

    # ── Compute overall score ──
    ds = sections.get('design_standards', {})
    ds_summary = ds.get('summary', {})
    # Exclude "Not Auditable" from score denominator
    ds_not_auditable = ds_summary.get('Not Auditable', 0)
    ds_total = sum(ds_summary.values()) - ds_not_auditable
    ds_met = ds_summary.get('Met', 0)
    ds_score = round((ds_met / ds_total * 100)) if ds_total > 0 else 0

    qa = sections.get('qa_categories', {})
    qa_summary = qa.get('summary', {})
    qa_total = sum(qa_summary.values())
    qa_pass = qa_summary.get('Pass', 0)
    qa_score = round((qa_pass / qa_total * 100)) if qa_total > 0 else 0

    a11y = sections.get('accessibility', {})
    a11y_summary = a11y.get('summary', {})
    a11y_critical = a11y_summary.get('Critical', 0)

    readiness = sections.get('readiness', {})
    readiness_cats = readiness.get('categories', [])
    readiness_pass = sum(1 for c in readiness_cats if c.get('status') == 'Pass')
    readiness_total = len(readiness_cats)
    # Determine status: explicit 'overall' field, derive from categories, or "Not Evaluated"
    if readiness.get('overall'):
        readiness_status = readiness['overall']
    elif readiness_total > 0:
        readiness_status = 'READY' if readiness_pass == readiness_total else 'NOT READY'
    else:
        readiness_status = 'Not Evaluated'

    # Use split scores from evaluator if available, else compute legacy
    readiness_score = data.get('readiness_score')
    design_score_val = data.get('design_score')
    a11y_score_val = data.get('a11y_score')
    if readiness_score is not None:
        overall_score = data.get('overall_score', readiness_score)
    else:
        overall_score = round((ds_score * 0.4 + qa_score * 0.3 + (100 if a11y_critical == 0 else max(0, 100 - a11y_critical * 25)) * 0.15 + (readiness_pass / readiness_total * 100 if readiness_total else 0) * 0.15))

    # Detect empty audit — all section totals are zero
    a11y_total = sum(a11y_summary.values())
    all_empty = (ds_total == 0 and qa_total == 0 and a11y_total == 0 and readiness_total == 0)
    empty_warning_html = ''
    if all_empty:
        import logging as _logging
        _logging.getLogger(__name__).warning(
            "Audit report generated with 0 findings across all sections. "
            "This usually means the course has no content, the wrong course ID is set, "
            "or audit data was not passed correctly. Verify Canvas connection and re-run."
        )
        empty_warning_html = (
            '<div style="background:#f59e0b;color:#000;padding:16px 24px;border-radius:8px;'
            'margin-bottom:24px;border-left:6px solid #b45309;">'
            '<strong>⚠️ Audit Warning: No findings generated.</strong> '
            'This usually means the course has no content, the course ID in <code>.env</code> '
            'is wrong, or audit data was not passed correctly. '
            'Verify your Canvas connection and re-run the audit.'
            '</div>'
        )

    # ── Per-Criterion Counts (for time estimate + flag-for-review) ──
    auto_confirmed = 0
    needs_review = 0
    criteria_total = 0
    for item in ds.get('items', []):
        criteria_results = item.get('criteria_results', [])
        if criteria_results:
            for cr in criteria_results:
                criteria_total += 1
                if cr.get('check_type') == 'deterministic' and cr.get('status') == 'Met':
                    auto_confirmed += 1
                else:
                    needs_review += 1
        else:
            # Legacy: no per-criterion data — count each standard as 1 review item
            criteria_total += 1
            needs_review += 1
    estimated_minutes = max(5, needs_review * 2)  # ~2 min per review item, minimum 5
    review_estimate_html = (
        f'<div style="text-align:center;padding:12px 0;color:#555;font-size:14px">'
        f'<strong>{auto_confirmed}</strong> of {criteria_total} criteria confirmed automatically. '
        f'<strong>{needs_review}</strong> item{"s" if needs_review != 1 else ""} '
        f'need{"s" if needs_review == 1 else ""} your review '
        f'(~{estimated_minutes} min estimated)</div>'
    ) if criteria_total > 0 else ''

    # ── Separate design standards into Confirmed vs Needs Review ──
    confirmed_items = []
    review_items = []
    for item in ds.get('items', []):
        # Item is "confirmed" only if ALL its criteria are deterministic+Met
        criteria_results = item.get('criteria_results', [])
        if criteria_results and all(
            cr.get('check_type') == 'deterministic' and cr.get('status') == 'Met'
            for cr in criteria_results
        ):
            confirmed_items.append(item)
        else:
            review_items.append(item)

    # ── Compute tier counts for filter bar ──
    tier_counts = {'design': 0, 'readiness': 0, 'a11y': 0}
    for item in ds.get('items', []):
        sid = str(item.get('id', ''))
        is_a11y = sid in ('22', '23') or str(item.get('criterion_id', '')).startswith(('22.', '23.'))
        if is_a11y:
            tier_counts['a11y'] += 1
        elif item.get('reviewer_tier') == 'id_assistant':
            tier_counts['readiness'] += 1
        else:
            tier_counts['design'] += 1

    # ── Design Standards Section ──
    ds_items_html = []
    for item in ds.get('items', []):
        status = item.get('status', 'Unknown')
        rec_html = ''
        if item.get('recommendation'):
            rec_html = f'<div class="finding-rec"><strong>Recommendation:</strong> {_escape(item["recommendation"])}</div>'
        # Confidence badge
        confidence = item.get('confidence', '')
        conf_colors = {'High': ('#1e7e34', '#e6f4ea'), 'Medium': ('#b5540a', '#fef3e5'), 'Low': ('#c62828', '#fde8e8')}
        conf_fg, conf_bg = conf_colors.get(confidence, ('#6a7883', '#f5f5f5'))
        conf_html = f'<span class="finding-badge" style="background:{conf_bg};color:{conf_fg};font-size:11px;margin-left:6px">{_escape(confidence)}</span>' if confidence else ''
        # Graph-verified badge (shown when evidence_source indicates graph)
        graph_badge = ''
        evidence_source = item.get('evidence_source', '')
        if 'graph' in str(evidence_source).lower():
            graph_badge = '<span class="finding-badge" style="background:#e8eaf6;color:#3949ab;font-size:11px;margin-left:6px">📊 Graph-verified</span>'
        elif evidence_source:
            graph_badge = f'<span class="finding-badge" style="background:#f5f5f5;color:#6a7883;font-size:11px;margin-left:6px">{_escape(evidence_source)}</span>'
        # Coverage badge
        coverage = item.get('coverage', '')
        coverage_html = f'<span style="color:#6a7883;font-size:11px;margin-left:8px">{_escape(coverage)}</span>' if coverage else ''
        # Determine tier category (used for both badge and filter)
        reviewer_tier = item.get('reviewer_tier', '')
        is_a11y = str(item.get('id', '')) in ('22', '23') or str(item.get('criterion_id', '')).startswith(('22.', '23.'))
        tier_filter = 'a11y' if is_a11y else ('readiness' if reviewer_tier == 'id_assistant' else 'design')
        # Reviewer tier badge — matches filter category
        tier_badges = {
            'design': '<span class="finding-badge" style="background:#f3ebe7;color:#AF674B;font-size:11px;margin-left:6px">Design</span>',
            'readiness': '<span class="finding-badge" style="background:#e0f4fa;color:#0081b3;font-size:11px;margin-left:6px">Readiness</span>',
            'a11y': '<span class="finding-badge" style="background:#e8eaf6;color:#3949ab;font-size:11px;margin-left:6px">A11y</span>',
        }
        tier_html = tier_badges.get(tier_filter, '')
        # Category badge (CRC vs Design Standard)
        cat_type = item.get('category_type', '')
        if not cat_type:
            cat_type = 'crc' if str(item.get('id', '')).startswith('crc') else ''
        cat_badge_html = '<span class="finding-badge" style="background:#f0f0f0;color:#6a7883;font-size:11px;margin-left:6px">CRC</span>' if cat_type == 'crc' else ''
        # Canvas link
        canvas_link = item.get('canvas_link', '')
        canvas_link_html = ''
        if canvas_link:
            canvas_link_html = f'<div style="margin-top:6px"><a href="{_escape(canvas_link)}" target="_blank" rel="noopener" style="color:#0081b3;font-size:12px;text-decoration:none">View in Canvas ↗</a></div>'
        ds_items_html.append(f'''
        <div class="finding-card" data-status="{status.lower().replace(' ', '-')}" data-tier="{tier_filter}">
          <div class="finding-header">
            <div class="finding-id-status">
              <span class="finding-id">{_escape(item.get("id", ""))}</span>
              <span class="finding-badge" style="background:{_severity_bg(status)};color:{_severity_color(status)}">{_severity_icon(status)} {_escape(status)}</span>
              {conf_html}{tier_html}{cat_badge_html}{graph_badge}{coverage_html}
            </div>
            <div class="finding-name">{_escape(item.get("name", ""))}</div>
            <div class="finding-category">{_escape(item.get("category", ""))}</div>
          </div>
          <div class="finding-body">
            <div class="finding-evidence"><strong>Evidence:</strong> {_escape(item.get("evidence", ""))}</div>
            {canvas_link_html}
            {rec_html}
            {_render_criteria_results(item.get("criteria_results", []))}
          </div>
        </div>''')

    # ── QA Categories Section ──
    qa_items_html = []
    for item in qa.get('items', []):
        status = item.get('status', 'Unknown')
        qa_tier = item.get('reviewer_tier', 'id_assistant')
        tier_label = 'Readiness' if qa_tier == 'id_assistant' else 'Design'
        tier_color = '#0081b3' if qa_tier == 'id_assistant' else '#AF674B'
        tier_bg = '#e0f4fa' if qa_tier == 'id_assistant' else '#f3ebe7'
        qa_items_html.append(f'''
        <tr>
          <td><span class="finding-badge" style="background:{_severity_bg(status)};color:{_severity_color(status)}">{_severity_icon(status)} {_escape(status)}</span></td>
          <td><strong>{_escape(item.get("id", ""))}</strong></td>
          <td>{_escape(item.get("name", ""))}</td>
          <td><span style="background:{tier_bg};color:{tier_color};font-size:11px;padding:2px 6px;border-radius:4px">{tier_label}</span></td>
          <td>{_escape(item.get("detail", ""))}</td>
        </tr>''')

    # ── Accessibility Section ──
    a11y_items_html = []
    for item in a11y.get('items', []):
        sev = item.get('severity', 'Info')
        a11y_link = item.get('canvas_link', '')
        page_name = _escape(item.get("page", ""))
        page_html = f'<a href="{_escape(a11y_link)}" target="_blank" rel="noopener" style="color:#0081b3;text-decoration:none"><code>{page_name}</code> ↗</a>' if a11y_link else f'<code>{page_name}</code>'
        a11y_items_html.append(f'''
        <tr>
          <td><span class="finding-badge" style="background:{_severity_bg(sev)};color:{_severity_color(sev)}">{_severity_icon(sev)} {_escape(sev)}</span></td>
          <td>{page_html}</td>
          <td>{_escape(item.get("issue", ""))}</td>
          <td><code style="font-size:11px">{_escape(item.get("element", ""))}</code></td>
          <td>{_escape(item.get("fix", ""))}</td>
        </tr>''')

    # ── Readiness Section ──
    readiness_html = []
    for cat in readiness_cats:
        cat_status = cat.get('status', 'Unknown')
        checks_html = []
        for check in cat.get('checks', []):
            cs = check.get('status', 'Unknown')
            note = f' — <em>{_escape(check.get("note", ""))}</em>' if check.get('note') else ''
            checks_html.append(f'''
            <div class="readiness-check">
              <span style="color:{_severity_color(cs)};font-weight:700">{_severity_icon(cs)}</span>
              {_escape(check.get("item", ""))}{note}
            </div>''')

        readiness_html.append(f'''
        <div class="readiness-category">
          <div class="readiness-cat-header">
            <span class="finding-badge" style="background:{_severity_bg(cat_status)};color:{_severity_color(cat_status)}">{_severity_icon(cat_status)} {_escape(cat_status)}</span>
            <strong>{_escape(cat.get("name", ""))}</strong>
          </div>
          <div class="readiness-checks">
            {''.join(checks_html)}
          </div>
        </div>''')

    # ── CLO Alignment ──
    clo_html = []
    for clo in clo_data.get('clos', []):
        mods = ', '.join(f'M{m}' for m in clo.get('modules', []))
        clo_html.append(f'''
        <tr>
          <td><strong>{_escape(clo.get("id", ""))}</strong></td>
          <td>{_escape(clo.get("text", ""))}</td>
          <td>{mods}</td>
          <td style="text-align:center">{clo.get("assessments", 0)}</td>
        </tr>''')

    # ── Remediation Roadmap ──
    remediation_section_html = _build_remediation_html(sections, overall_score)

    # ── Score ring color ──
    score_color = '#1e7e34' if overall_score >= 80 else '#b5540a' if overall_score >= 60 else '#c62828'

    return f'''<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Course Audit Report — {_escape(course.get("name", "Course"))}</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link href="https://fonts.googleapis.com/css2?family=Roboto:wght@300;400;500;700&family=JetBrains+Mono:wght@400&display=swap" rel="stylesheet">
  <style>
    :root {{
      --canvas-text: #2d3b45;
      --canvas-text-secondary: #6a7883;
      --canvas-link: #0374b5;
      --canvas-border: #c7cdd1;
      --asu-maroon: #8C1D40;
      --asu-gold: #FFC627;
    }}
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{
      font-family: 'Roboto', 'Helvetica Neue', sans-serif;
      font-size: 15px;
      line-height: 1.6;
      color: var(--canvas-text);
      background: #f0f0f0;
    }}

    /* ── Header ── */
    .report-header {{
      background: linear-gradient(135deg, var(--asu-maroon) 0%, #6a1530 100%);
      color: white;
      padding: 40px 48px;
    }}
    .report-header-top {{
      display: flex;
      justify-content: space-between;
      align-items: flex-start;
    }}
    .report-header h1 {{
      font-size: 28px;
      font-weight: 400;
      margin-bottom: 4px;
    }}
    .report-header h1 strong {{ font-weight: 700; }}
    .report-header .subtitle {{
      font-size: 14px;
      opacity: 0.8;
    }}
    .report-meta {{
      display: flex;
      gap: 24px;
      margin-top: 16px;
      font-size: 13px;
      opacity: 0.85;
    }}

    /* ── Score Ring ── */
    .score-ring {{
      width: 100px;
      height: 100px;
      position: relative;
    }}
    .score-ring svg {{
      transform: rotate(-90deg);
    }}
    .score-ring-value {{
      position: absolute;
      top: 50%;
      left: 50%;
      transform: translate(-50%, -50%);
      font-size: 28px;
      font-weight: 700;
      color: white;
    }}

    /* ── Container ── */
    .container {{
      max-width: 1100px;
      margin: 0 auto;
      padding: 32px 24px 80px;
    }}

    /* ── Summary Cards ── */
    .summary-grid {{
      display: grid;
      grid-template-columns: repeat(4, 1fr);
      gap: 16px;
      margin: -32px 0 32px;
      position: relative;
      z-index: 10;
    }}
    .summary-card {{
      background: white;
      border-radius: 8px;
      padding: 20px 24px;
      box-shadow: 0 2px 8px rgba(0,0,0,0.08);
      border-top: 3px solid var(--canvas-border);
    }}
    .summary-card h3 {{
      font-size: 12px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.06em;
      color: var(--canvas-text-secondary);
      margin-bottom: 12px;
    }}
    .summary-card .summary-stats {{
      display: flex;
      gap: 12px;
      flex-wrap: wrap;
    }}
    .summary-stat {{
      display: flex;
      align-items: center;
      gap: 6px;
      font-size: 13px;
    }}
    .summary-stat-num {{
      font-size: 22px;
      font-weight: 700;
    }}

    /* ── Section ── */
    .report-section {{
      background: white;
      border-radius: 8px;
      box-shadow: 0 1px 4px rgba(0,0,0,0.06);
      margin-bottom: 24px;
      overflow: hidden;
    }}
    .section-header {{
      padding: 20px 24px;
      border-bottom: 1px solid #eee;
      display: flex;
      justify-content: space-between;
      align-items: center;
      cursor: pointer;
      user-select: none;
    }}
    .section-header:hover {{ background: #fafbfc; }}
    .section-header h2 {{
      font-size: 18px;
      font-weight: 500;
    }}
    .section-header .section-subtitle {{
      font-size: 13px;
      color: var(--canvas-text-secondary);
    }}
    .section-header .section-toggle {{
      font-size: 20px;
      color: var(--canvas-text-secondary);
      transition: transform 0.2s;
    }}
    .report-section.collapsed .section-body {{ display: none; }}
    .report-section.collapsed .section-toggle {{ transform: rotate(-90deg); }}
    .section-body {{ padding: 0; }}

    /* ── Filter Bar ── */
    .filter-bar {{
      padding: 12px 24px;
      background: #fafbfc;
      border-bottom: 1px solid #eee;
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
    }}
    .filter-btn {{
      padding: 4px 12px;
      border-radius: 20px;
      border: 1px solid var(--canvas-border);
      background: white;
      font-size: 12px;
      cursor: pointer;
      font-family: inherit;
    }}
    .filter-btn:hover {{ background: #f0f4ff; }}
    .filter-btn.active {{ background: var(--canvas-link); color: white; border-color: var(--canvas-link); }}
    .tier-btn {{
      padding: 3px 10px;
      border-radius: 20px;
      font-size: 11px;
      cursor: pointer;
      font-family: inherit;
      font-weight: 600;
      transition: opacity 0.2s;
    }}
    .tier-btn:not(.active) {{ opacity: 0.4; }}

    /* ── Finding Card ── */
    .finding-card {{
      padding: 16px 24px;
      border-bottom: 1px solid #f0f0f0;
    }}
    .finding-card:last-child {{ border-bottom: none; }}
    .finding-header {{ margin-bottom: 8px; }}
    .finding-id-status {{
      display: flex;
      align-items: center;
      gap: 10px;
      margin-bottom: 4px;
    }}
    .finding-id {{
      font-family: 'JetBrains Mono', monospace;
      font-size: 12px;
      color: var(--canvas-text-secondary);
      background: #f5f5f5;
      padding: 1px 8px;
      border-radius: 3px;
    }}
    .finding-badge {{
      display: inline-flex;
      align-items: center;
      gap: 4px;
      padding: 2px 10px;
      border-radius: 12px;
      font-size: 12px;
      font-weight: 600;
      white-space: nowrap;
    }}
    .finding-name {{
      font-size: 15px;
      font-weight: 500;
    }}
    .finding-category {{
      font-size: 12px;
      color: var(--canvas-text-secondary);
    }}
    .finding-body {{
      font-size: 14px;
      color: #444;
      line-height: 1.5;
    }}
    .finding-evidence {{ margin-bottom: 6px; }}
    .finding-rec {{
      background: #fff8e1;
      padding: 8px 12px;
      border-radius: 4px;
      border-left: 3px solid var(--asu-gold);
      font-size: 13px;
    }}

    /* ── Tables ── */
    .report-table {{
      width: 100%;
      border-collapse: collapse;
    }}
    .report-table th, .report-table td {{
      padding: 10px 16px;
      text-align: left;
      border-bottom: 1px solid #f0f0f0;
      font-size: 13px;
    }}
    .report-table th {{
      background: #fafbfc;
      font-weight: 600;
      font-size: 12px;
      text-transform: uppercase;
      letter-spacing: 0.04em;
      color: var(--canvas-text-secondary);
      position: sticky;
      top: 0;
    }}
    .report-table code {{
      background: #f5f5f5;
      padding: 1px 6px;
      border-radius: 3px;
      font-size: 12px;
    }}

    /* ── Readiness ── */
    .readiness-category {{
      padding: 16px 24px;
      border-bottom: 1px solid #f0f0f0;
    }}
    .readiness-cat-header {{
      display: flex;
      align-items: center;
      gap: 10px;
      margin-bottom: 8px;
    }}
    .readiness-checks {{
      padding-left: 16px;
    }}
    .readiness-check {{
      display: flex;
      align-items: baseline;
      gap: 8px;
      font-size: 13px;
      padding: 3px 0;
    }}

    /* ── CLO table ── */
    .clo-modules {{
      display: flex;
      gap: 4px;
      flex-wrap: wrap;
    }}
    .clo-mod {{
      background: var(--asu-maroon);
      color: white;
      font-size: 10px;
      padding: 1px 6px;
      border-radius: 3px;
      font-weight: 600;
    }}

    /* ── Print ── */
    @media print {{
      body {{ background: white; }}
      .report-header {{ padding: 24px; }}
      .container {{ padding: 16px 0; }}
      .summary-grid {{ margin-top: 16px; }}
      .report-section {{ box-shadow: none; border: 1px solid #ddd; break-inside: avoid; }}
      .report-section.collapsed .section-body {{ display: block !important; }}
      .filter-bar {{ display: none; }}
    }}

    @media (max-width: 768px) {{
      .summary-grid {{ grid-template-columns: repeat(2, 1fr); }}
      .report-header {{ padding: 24px; }}
      .report-header h1 {{ font-size: 22px; }}
    }}
  </style>
</head>
<body>

  <!-- Report Header -->
  <header class="report-header">
    <div class="report-header-top">
      <div>
        <h1><strong>Course Audit Report</strong></h1>
        <div class="subtitle">{_escape(course.get('name', 'Course'))}</div>
        <div class="report-meta">
          <span>Course ID: {_escape(course.get('id', ''))}</span>
          <span>{_escape(course.get('institution', ''))}</span>
          <span>{date_display}</span>
          <span>Generated by {_escape(auditor)}</span>
        </div>
      </div>
      <div class="score-ring" title="Overall Score: {overall_score}%">
        <svg width="100" height="100" viewBox="0 0 100 100">
          <circle cx="50" cy="50" r="42" fill="none" stroke="rgba(255,255,255,0.2)" stroke-width="8"/>
          <circle cx="50" cy="50" r="42" fill="none" stroke="{score_color}" stroke-width="8"
            stroke-dasharray="{round(264 * overall_score / 100)} 264"
            stroke-linecap="round"/>
        </svg>
        <div class="score-ring-value">{overall_score}</div>
      </div>
    </div>
  </header>

  <div class="container">

    {empty_warning_html}

    <!-- Score Breakdown -->
    <div style="display:flex;gap:16px;justify-content:center;margin-bottom:12px;flex-wrap:wrap">
      <div style="text-align:center;padding:12px 24px;background:white;border-radius:8px;border:1px solid #eee">
        <div style="font-size:28px;font-weight:700;color:{'#1e7e34' if (readiness_score or 0) >= 80 else '#b5540a' if (readiness_score or 0) >= 50 else '#c62828'}">{readiness_score if readiness_score is not None else '—'}%</div>
        <div style="font-size:12px;color:#888;font-weight:500">READINESS (Col B)</div>
      </div>
      <div style="text-align:center;padding:12px 24px;background:white;border-radius:8px;border:1px solid #eee">
        <div style="font-size:28px;font-weight:700;color:{'#1e7e34' if (design_score_val or 0) >= 80 else '#b5540a' if (design_score_val or 0) >= 50 else '#6a7883' if design_score_val is None else '#c62828'}">{f'{design_score_val}%' if design_score_val is not None else 'N/A'}</div>
        <div style="font-size:12px;color:#888;font-weight:500">DESIGN (Col C)</div>
      </div>
      <div style="text-align:center;padding:12px 24px;background:white;border-radius:8px;border:1px solid {'#c62828' if (a11y_score_val or 0) < 50 else '#eee'}">
        <div style="font-size:28px;font-weight:700;color:{'#1e7e34' if (a11y_score_val or 0) >= 80 else '#b5540a' if (a11y_score_val or 0) >= 50 else '#c62828'}">{a11y_score_val if a11y_score_val is not None else '—'}%</div>
        <div style="font-size:12px;color:#888;font-weight:500">A11Y (ASU Mandated)</div>
      </div>
    </div>

    <!-- Summary Cards -->
    <div class="summary-grid">
      <div class="summary-card" style="border-top-color:#1e7e34">
        <h3>Design Standards</h3>
        <div class="summary-stats">
          <div class="summary-stat"><span class="summary-stat-num" style="color:#1e7e34">{ds_summary.get('Met', 0)}</span> Met</div>
          <div class="summary-stat"><span class="summary-stat-num" style="color:#b5540a">{ds_summary.get('Partially Met', 0)}</span> Partial</div>
          <div class="summary-stat"><span class="summary-stat-num" style="color:#c62828">{ds_summary.get('Not Met', 0)}</span> Not Met</div>
          {f'<div class="summary-stat"><span class="summary-stat-num" style="color:#6a7883">{ds_summary.get("Not Auditable", 0)}</span> N/A</div>' if ds_summary.get('Not Auditable', 0) > 0 else ''}
        </div>
      </div>
      <div class="summary-card" style="border-top-color:#0374b5">
        <h3>QA Categories</h3>
        <div class="summary-stats">
          <div class="summary-stat"><span class="summary-stat-num" style="color:#1e7e34">{qa_summary.get('Pass', 0)}</span> Pass</div>
          <div class="summary-stat"><span class="summary-stat-num" style="color:#b5540a">{qa_summary.get('Warn', 0)}</span> Warn</div>
          <div class="summary-stat"><span class="summary-stat-num" style="color:#c62828">{qa_summary.get('Fail', 0)}</span> Fail</div>
        </div>
      </div>
      <div class="summary-card" style="border-top-color:#7b2d8b">
        <h3>Accessibility</h3>
        <div class="summary-stats">
          <div class="summary-stat"><span class="summary-stat-num" style="color:#c62828">{a11y_summary.get('Critical', 0)}</span> Critical</div>
          <div class="summary-stat"><span class="summary-stat-num" style="color:#b5540a">{a11y_summary.get('Warning', 0)}</span> Warning</div>
          <div class="summary-stat"><span class="summary-stat-num" style="color:#1e7e34">{a11y_summary.get('Info', 0)}</span> Info</div>
        </div>
      </div>
      <div class="summary-card" style="border-top-color:{'#1e7e34' if readiness_status == 'READY' else '#6a7883' if readiness_status == 'Not Evaluated' else '#c62828'}">
        <h3>Course Readiness</h3>
        <div class="summary-stats">
          <div class="summary-stat"><span class="summary-stat-num" style="color:{'#1e7e34' if readiness_status == 'READY' else '#6a7883' if readiness_status == 'Not Evaluated' else '#c62828'}">{readiness_status}</span></div>
          <div class="summary-stat"><span class="summary-stat-num">{readiness_pass}</span>/{readiness_total}</div>
        </div>
      </div>
    </div>

    <!-- Design Standards Section -->
    <div class="report-section" id="section-standards">
      <div class="section-header" onclick="this.parentElement.classList.toggle('collapsed')">
        <div>
          <h2>{_escape(ds.get('title', 'Design Standards'))}</h2>
          <div class="section-subtitle">{_escape(ds.get('subtitle', ''))}</div>
        </div>
        <span class="section-toggle">▾</span>
      </div>
      <div class="filter-bar">
        <button class="filter-btn active" onclick="filterFindings(this, 'standards', 'all')">All ({ds_total})</button>
        <button class="filter-btn" onclick="filterFindings(this, 'standards', 'met')">Met ({ds_summary.get('Met', 0)})</button>
        <button class="filter-btn" onclick="filterFindings(this, 'standards', 'partially-met')">Partially Met ({ds_summary.get('Partially Met', 0)})</button>
        <button class="filter-btn" onclick="filterFindings(this, 'standards', 'not-met')">Not Met ({ds_summary.get('Not Met', 0)})</button>
        {f'<button class="filter-btn" onclick="filterFindings(this, \'standards\', \'not-auditable\')">Not Auditable ({ds_summary.get("Not Auditable", 0)})</button>' if ds_summary.get('Not Auditable', 0) > 0 else ''}
      </div>
      <div class="filter-bar" style="margin-top:4px">
        <span style="font-size:11px;color:#888;margin-right:8px">Category:</span>
        <button class="tier-btn active" onclick="toggleTier(this, 'standards', 'design')" style="background:#f3ebe7;color:#AF674B;border:1px solid #AF674B33">Design ({tier_counts['design']})</button>
        <button class="tier-btn active" onclick="toggleTier(this, 'standards', 'readiness')" style="background:#e0f4fa;color:#0081b3;border:1px solid #0081b333">Readiness ({tier_counts['readiness']})</button>
        <button class="tier-btn active" onclick="toggleTier(this, 'standards', 'a11y')" style="background:#e8eaf6;color:#3949ab;border:1px solid #3949ab33">A11y ({tier_counts['a11y']})</button>
      </div>
      {review_estimate_html}
      <div class="section-body" id="standards-body">
        {''.join(ds_items_html)}
      </div>
    </div>

    <!-- QA Categories Section -->
    <div class="report-section" id="section-qa">
      <div class="section-header" onclick="this.parentElement.classList.toggle('collapsed')">
        <div>
          <h2>{_escape(qa.get('title', 'QA Categories'))}</h2>
          <div class="section-subtitle">{_escape(qa.get('subtitle', ''))}</div>
        </div>
        <span class="section-toggle">▾</span>
      </div>
      <div class="section-body">
        <table class="report-table">
          <thead>
            <tr><th>Status</th><th>ID</th><th>Category</th><th>Tier</th><th>Detail</th></tr>
          </thead>
          <tbody>
            {''.join(qa_items_html)}
          </tbody>
        </table>
      </div>
    </div>

    <!-- Accessibility Section -->
    <div class="report-section" id="section-a11y">
      <div class="section-header" onclick="this.parentElement.classList.toggle('collapsed')">
        <div>
          <h2>{_escape(a11y.get('title', 'Accessibility'))}</h2>
          <div class="section-subtitle">{_escape(a11y.get('subtitle', ''))}</div>
        </div>
        <span class="section-toggle">▾</span>
      </div>
      <div class="section-body">
        <table class="report-table">
          <thead>
            <tr><th>Severity</th><th>Page</th><th>Issue</th><th>Element</th><th>Suggested Fix</th></tr>
          </thead>
          <tbody>
            {''.join(a11y_items_html)}
          </tbody>
        </table>
      </div>
    </div>

    <!-- Readiness Section -->
    <div class="report-section" id="section-readiness">
      <div class="section-header" onclick="this.parentElement.classList.toggle('collapsed')">
        <div>
          <h2>{_escape(readiness.get('title', 'Course Readiness'))}</h2>
          <div class="section-subtitle">{_escape(readiness.get('subtitle', ''))}</div>
        </div>
        <span class="section-toggle">▾</span>
      </div>
      <div class="section-body">
        {''.join(readiness_html)}
      </div>
    </div>

    {remediation_section_html}

    <!-- CLO Alignment Matrix -->
    <div class="report-section" id="section-clo">
      <div class="section-header" onclick="this.parentElement.classList.toggle('collapsed')">
        <div>
          <h2>{_escape(clo_data.get('title', 'CLO Alignment Matrix'))}</h2>
          <div class="section-subtitle">Course Learning Objectives mapped to modules and assessments</div>
        </div>
        <span class="section-toggle">▾</span>
      </div>
      <div class="section-body">
        <table class="report-table">
          <thead>
            <tr><th>CLO</th><th>Objective</th><th>Modules</th><th>Assessments</th></tr>
          </thead>
          <tbody>
            {''.join(clo_html)}
          </tbody>
        </table>
      </div>
    </div>

    <!-- Footer -->
    <div style="text-align:center;padding:24px;color:var(--canvas-text-secondary);font-size:12px">
      Generated by ID Workbench · {date_display} · <a href="https://github.com/bespined/ID-Workbench" style="color:var(--canvas-link)">github.com/bespined/ID-Workbench</a>
    </div>

  </div>

  <script>
    function filterFindings(btn, section, status) {{
      // Update active button in status bar
      btn.closest('.filter-bar').querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
      btn.classList.add('active');
      applyFilters(section);
    }}

    function toggleTier(btn, section, tier) {{
      // Toggle this tier button on/off
      btn.classList.toggle('active');
      if (!btn.classList.contains('active')) {{
        btn.style.opacity = '0.4';
      }} else {{
        btn.style.opacity = '1';
      }}
      applyFilters(section);
    }}

    function applyFilters(section) {{
      const body = document.getElementById(section + '-body');
      if (!body) return;

      // Get active status filter
      const statusBar = body.closest('.report-section').querySelector('.filter-bar');
      const activeStatus = statusBar ? statusBar.querySelector('.filter-btn.active') : null;
      const status = activeStatus ? activeStatus.textContent.split(' (')[0].toLowerCase().replace(' ', '-') : 'all';

      // Get active tier filters
      const tierBtns = body.closest('.report-section').querySelectorAll('.tier-btn');
      const activeTiers = new Set();
      tierBtns.forEach(b => {{
        if (b.classList.contains('active')) {{
          // Extract tier from onclick attribute
          const match = b.getAttribute('onclick').match(/toggleTier\\(this,\\s*'[^']+',\\s*'([^']+)'\\)/);
          if (match) activeTiers.add(match[1]);
        }}
      }});

      body.querySelectorAll('.finding-card').forEach(card => {{
        const statusMatch = status === 'all' || card.dataset.status === status;
        const tierMatch = activeTiers.size === 0 || activeTiers.has(card.dataset.tier || '');
        card.style.display = (statusMatch && tierMatch) ? '' : 'none';
      }});
    }}
  </script>

</body>
</html>'''


# ============================================================
# XLSX REPORT GENERATION (QA Initiate Format)
# ============================================================

TEMPLATE_PATH = PLUGIN_ROOT / "templates" / "[Final Template] QA Initiate.xlsx"

# Map standard IDs to their formula row and criteria row range in the template.
# Format: standard_id -> (formula_row, first_criteria_row, last_criteria_row)
STANDARD_ROW_MAP = {
    "01": (4, 4, 7),    "02": (8, 8, 10),   "03": (11, 11, 12),
    "04": (13, 13, 19),  "05": (20, 20, 23),  "06": (24, 24, 29),
    "07": (30, 30, 31),  "08": (34, 34, 37),  "09": (38, 38, 41),
    "10": (42, 42, 44),  "11": (45, 45, 48),  "12": (51, 51, 53),
    "13": (54, 54, 55),  "14": (56, 56, 58),  "15": (59, 59, 60),
    "16": (61, 61, 62),  "17": (65, 65, 68),  "18": (69, 69, 73),
    "19": (74, 74, 76),  "20": (77, 77, 78),  "21": (81, 81, 83),
    "22": (84, 84, 85),  "23": (86, 86, 87),  "24": (88, 88, 90),
    "25": (91, 91, 92),
}

# All criteria rows (column B is user-fillable)
ALL_CRITERIA_ROWS = [
    4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,
    34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,
    51,52,53,54,55,56,57,58,59,60,61,62,
    65,66,67,68,69,70,71,72,73,74,75,76,77,78,
    81,82,83,84,85,86,87,88,89,90,91,92
]


def _status_fill(status):
    """Return PatternFill for a status value."""
    from openpyxl.styles import PatternFill
    s = status.lower() if status else ""
    if s == "met":
        return PatternFill("solid", fgColor="E6F4EA")
    elif s == "partially met":
        return PatternFill("solid", fgColor="FEF3E5")
    elif s == "not met":
        return PatternFill("solid", fgColor="FDE8E8")
    return PatternFill()


def _status_font(status):
    """Return Font for a status value."""
    from openpyxl.styles import Font
    s = status.lower() if status else ""
    if s == "met":
        return Font(color="1E7E34")
    elif s == "partially met":
        return Font(color="B5540A")
    elif s == "not met":
        return Font(color="C62828", bold=True)
    return Font()


def _confidence_font(conf):
    """Return Font for a confidence value."""
    from openpyxl.styles import Font
    c = conf.lower() if conf else ""
    if c == "high":
        return Font(color="1E7E34")
    elif c == "medium":
        return Font(color="B5540A")
    elif c == "low":
        return Font(color="C62828", bold=True)
    return Font()


def generate_xlsx_report(data: dict, output_path: str) -> dict:
    """Generate XLSX audit report in QA Initiate format.

    Loads the QA Initiate template, populates it with audit results,
    adds a Dashboard sheet and External Links sheet.

    Returns dict with path, stats.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter

    if not TEMPLATE_PATH.exists():
        _log.error(f"Template not found: {TEMPLATE_PATH}")
        return {"ok": False, "error": "QA Initiate template not found"}

    wb = load_workbook(str(TEMPLATE_PATH))
    ws = wb.active  # "QA Initiate" sheet

    course = data.get("course", {})
    audit_date = data.get("audit_date", datetime.now().isoformat())
    sections = data.get("sections", {})

    try:
        date_display = datetime.fromisoformat(audit_date).strftime("%Y-%m-%d %H:%M")
    except (ValueError, TypeError):
        date_display = str(audit_date)

    # ── Set course title ──
    ws["A1"] = f"{course.get('name', 'Unknown Course')} (ID: {course.get('id', 'N/A')}) — Audit {date_display}"

    # ── Populate design standards criteria ──
    ds = sections.get("design_standards", {})
    ds_items = ds.get("items", [])

    # Build lookup by standard ID
    items_by_id = {}
    for item in ds_items:
        sid = str(item.get("id", "")).zfill(2)
        items_by_id[sid] = item

    stats = {"met": 0, "partially_met": 0, "not_met": 0, "filled": 0}

    for sid, (formula_row, first_row, last_row) in STANDARD_ROW_MAP.items():
        item = items_by_id.get(sid)
        if not item:
            continue

        status = item.get("status", "Not Met")
        evidence = item.get("evidence", "")
        recommendation = item.get("recommendation", "")
        confidence = item.get("confidence", "Medium")
        coverage = item.get("coverage", "")
        scope = item.get("scope", "Course-wide")
        evidence_source = item.get("evidence_source", "Canvas")
        canvas_link = item.get("canvas_link", "")

        # Count stats
        sl = status.lower()
        if sl == "met":
            stats["met"] += 1
        elif sl == "partially met":
            stats["partially_met"] += 1
        else:
            stats["not_met"] += 1

        # Fill the first criteria row for this standard
        target_row = first_row

        # Column B: Status
        cell_b = ws.cell(row=target_row, column=2)
        cell_b.value = status
        cell_b.fill = _status_fill(status)
        cell_b.font = _status_font(status)

        # Also fill remaining criteria rows in this standard with same status
        for r in range(first_row + 1, last_row + 1):
            cb = ws.cell(row=r, column=2)
            cb.value = status
            cb.fill = _status_fill(status)
            cb.font = _status_font(status)

        # Column D: Evidence / Reviewer Notes
        ws.cell(row=target_row, column=4).value = evidence

        # Column E: Recommendations
        if recommendation:
            ws.cell(row=target_row, column=5).value = recommendation

        # Column F: Confidence
        cell_f = ws.cell(row=target_row, column=6)
        cell_f.value = confidence
        cell_f.font = _confidence_font(confidence)
        cell_f.alignment = Alignment(horizontal="center")

        # Column G: Coverage
        ws.cell(row=target_row, column=7).value = coverage
        ws.cell(row=target_row, column=7).alignment = Alignment(horizontal="center")

        # Column H: Scope
        ws.cell(row=target_row, column=8).value = scope
        ws.cell(row=target_row, column=8).alignment = Alignment(horizontal="center")

        # Column I: Evidence Source
        ws.cell(row=target_row, column=9).value = evidence_source
        ws.cell(row=target_row, column=9).alignment = Alignment(horizontal="center")

        # Column J: Canvas Link
        if canvas_link:
            cell_j = ws.cell(row=target_row, column=10)
            cell_j.value = canvas_link
            cell_j.font = Font(color="0563C1", underline="single")
            cell_j.hyperlink = canvas_link

        # Column K: Reviewer Tier
        r_tier = item.get("reviewer_tier", "id")
        tier_label = "IDA" if r_tier == "id_assistant" else "ID"
        cell_k = ws.cell(row=target_row, column=11)
        cell_k.value = tier_label
        cell_k.alignment = Alignment(horizontal="center")
        if r_tier == "id_assistant":
            cell_k.font = Font(color="0081B3", bold=True)
        else:
            cell_k.font = Font(color="AF674B", bold=True)

        stats["filled"] += 1

    # ── Create Dashboard Sheet ──
    _create_dashboard_sheet(wb, data, stats)

    # ── Create External Links Sheet ──
    _create_external_links_sheet(wb, data)

    # ── Save ──
    output_p = Path(output_path)
    wb.save(str(output_p))
    _log.info(f"XLSX report saved to: {output_p}")

    # Recalculate formulas
    recalc_script = PLUGIN_ROOT / "scripts" / "recalc.py"
    if recalc_script.exists():
        try:
            result = subprocess.run(
                [sys.executable, str(recalc_script), str(output_p)],
                capture_output=True, text=True, timeout=60
            )
            if result.returncode == 0:
                _log.info("Formula recalculation complete")
            else:
                _log.warning(f"Recalc warning: {result.stderr[:200]}")
        except Exception as e:
            _log.warning(f"Recalc skipped: {e}")

    return {
        "ok": True,
        "path": str(output_p),
        "standards_filled": stats["filled"],
        "met": stats["met"],
        "partially_met": stats["partially_met"],
        "not_met": stats["not_met"],
    }


def _create_dashboard_sheet(wb, data, standards_stats):
    """Create the Dashboard summary sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter

    ds = wb.create_sheet("Dashboard", 0)  # Insert as first sheet

    # Colors
    maroon = "8C1D40"
    gold = "FFC627"
    green = "1E7E34"
    amber = "B5540A"
    red = "C62828"
    light_gray = "F5F5F5"

    hdr_font = Font(name="Arial", size=20, bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor=maroon)
    section_font = Font(name="Arial", size=14, bold=True, color=maroon)
    label_font = Font(name="Arial", size=11)
    value_font = Font(name="Arial", size=14, bold=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    course = data.get("course", {})
    audit_date = data.get("audit_date", "")
    sections = data.get("sections", {})

    try:
        date_display = datetime.fromisoformat(audit_date).strftime("%B %d, %Y")
    except (ValueError, TypeError):
        date_display = str(audit_date)

    # ── Title Row ──
    ds.merge_cells("A1:H1")
    ds["A1"] = "Course Audit Dashboard"
    ds["A1"].font = hdr_font
    ds["A1"].fill = hdr_fill
    ds["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ds.row_dimensions[1].height = 40

    # ── Course Info ──
    ds.merge_cells("A2:H2")
    ds["A2"] = f"Course: {course.get('name', 'N/A')}  •  ID: {course.get('id', 'N/A')}  •  Date: {date_display}  •  Auditor: {_resolve_auditor(data)}"
    ds["A2"].font = Font(name="Arial", size=10, italic=True)
    ds["A2"].alignment = Alignment(horizontal="center")

    # ── Overall Score ──
    ds_summary = sections.get("design_standards", {}).get("summary", {})
    qa_summary = sections.get("qa_categories", {}).get("summary", {})
    a11y_summary = sections.get("accessibility", {}).get("summary", {})
    readiness = sections.get("readiness", {})
    readiness_cats = readiness.get("categories", [])

    ds_total = (sum(ds_summary.values()) - ds_summary.get("Not Auditable", 0)) or 1
    ds_met = ds_summary.get("Met", 0)
    ds_score = round(ds_met / ds_total * 100)

    qa_total = sum(qa_summary.values()) or 1
    qa_pass = qa_summary.get("Pass", 0)
    qa_score = round(qa_pass / qa_total * 100)

    a11y_critical = a11y_summary.get("Critical", 0)
    a11y_score = max(0, 100 - a11y_critical * 25)

    readiness_pass = sum(1 for c in readiness_cats if c.get("status") == "Pass")
    readiness_total = len(readiness_cats) or 1
    readiness_score = round(readiness_pass / readiness_total * 100)

    overall = round(ds_score * 0.4 + qa_score * 0.3 + a11y_score * 0.15 + readiness_score * 0.15)

    ds.merge_cells("A4:B4")
    ds["A4"] = "OVERALL SCORE"
    ds["A4"].font = section_font

    ds["C4"] = f"{overall}%"
    ds["C4"].font = Font(name="Arial", size=28, bold=True, color=green if overall >= 80 else (amber if overall >= 60 else red))

    ds["D4"] = f"(Standards {ds_score}% × 40%  +  QA {qa_score}% × 30%  +  A11y {a11y_score}% × 15%  +  Readiness {readiness_score}% × 15%)"
    ds["D4"].font = Font(name="Arial", size=9, italic=True, color="666666")

    # ── Reviewer Tier Breakdown ──
    ds_items = sections.get("design_standards", {}).get("items", [])
    ida_count = sum(1 for it in ds_items if it.get("reviewer_tier") == "id_assistant")
    id_count = len(ds_items) - ida_count
    ds["A5"] = "Reviewer Tiers:"
    ds["A5"].font = label_font
    ds["B5"] = f"IDA-reviewable: {ida_count}"
    ds["B5"].font = Font(name="Arial", size=11, color="0081B3", bold=True)
    ds["C5"] = f"ID-reviewable: {id_count}"
    ds["C5"].font = Font(name="Arial", size=11, color="AF674B", bold=True)

    # ── Design Standards Summary ──
    row = 7
    ds.cell(row=row, column=1, value="DESIGN STANDARDS (25)").font = section_font
    row += 1
    gray = "6A7883"
    for label, count, color in [
        ("Met", ds_summary.get("Met", 0), green),
        ("Partially Met", ds_summary.get("Partially Met", 0), amber),
        ("Not Met", ds_summary.get("Not Met", 0), red),
        ("Not Auditable", ds_summary.get("Not Auditable", 0), gray),
    ]:
        ds.cell(row=row, column=1, value=label).font = label_font
        c = ds.cell(row=row, column=2, value=count)
        c.font = Font(name="Arial", size=14, bold=True, color=color)
        c.alignment = Alignment(horizontal="center")
        row += 1

    # Chart data for standards
    chart_data_row = row
    ds.cell(row=chart_data_row, column=4, value="Met")
    ds.cell(row=chart_data_row, column=5, value="Partially Met")
    ds.cell(row=chart_data_row, column=6, value="Not Met")
    ds.cell(row=chart_data_row + 1, column=4, value=ds_summary.get("Met", 0))
    ds.cell(row=chart_data_row + 1, column=5, value=ds_summary.get("Partially Met", 0))
    ds.cell(row=chart_data_row + 1, column=6, value=ds_summary.get("Not Met", 0))

    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Design Standards"
    chart1.y_axis.title = "Count"
    chart1.style = 10
    chart1.width = 16
    chart1.height = 10
    cats = Reference(ds, min_col=4, max_col=6, min_row=chart_data_row)
    vals = Reference(ds, min_col=4, max_col=6, min_row=chart_data_row + 1)
    chart1.add_data(vals, from_rows=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    # Color the bars
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    if chart1.series:
        s = chart1.series[0]
        pt_met = DataPoint(idx=0)
        pt_met.graphicalProperties.solidFill = green
        s.data_points.append(pt_met)
        pt_partial = DataPoint(idx=1)
        pt_partial.graphicalProperties.solidFill = amber
        s.data_points.append(pt_partial)
        pt_not = DataPoint(idx=2)
        pt_not.graphicalProperties.solidFill = red
        s.data_points.append(pt_not)
    ds.add_chart(chart1, f"D{row - 3}")

    # ── QA Categories Summary ──
    row = chart_data_row + 3
    ds.cell(row=row, column=1, value="QA CATEGORIES (19)").font = section_font
    row += 1
    for label, count, color in [
        ("Pass", qa_summary.get("Pass", 0), green),
        ("Warn", qa_summary.get("Warn", 0), amber),
        ("Fail", qa_summary.get("Fail", 0), red),
    ]:
        ds.cell(row=row, column=1, value=label).font = label_font
        c = ds.cell(row=row, column=2, value=count)
        c.font = Font(name="Arial", size=14, bold=True, color=color)
        c.alignment = Alignment(horizontal="center")
        row += 1

    # QA chart data
    qa_chart_row = row
    ds.cell(row=qa_chart_row, column=4, value="Pass")
    ds.cell(row=qa_chart_row, column=5, value="Warn")
    ds.cell(row=qa_chart_row, column=6, value="Fail")
    ds.cell(row=qa_chart_row + 1, column=4, value=qa_summary.get("Pass", 0))
    ds.cell(row=qa_chart_row + 1, column=5, value=qa_summary.get("Warn", 0))
    ds.cell(row=qa_chart_row + 1, column=6, value=qa_summary.get("Fail", 0))

    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "QA Categories"
    chart2.y_axis.title = "Count"
    chart2.style = 10
    chart2.width = 16
    chart2.height = 10
    cats2 = Reference(ds, min_col=4, max_col=6, min_row=qa_chart_row)
    vals2 = Reference(ds, min_col=4, max_col=6, min_row=qa_chart_row + 1)
    chart2.add_data(vals2, from_rows=True)
    chart2.set_categories(cats2)
    if chart2.series:
        s2 = chart2.series[0]
        pt0 = DataPoint(idx=0)
        pt0.graphicalProperties.solidFill = green
        s2.data_points.append(pt0)
        pt1 = DataPoint(idx=1)
        pt1.graphicalProperties.solidFill = amber
        s2.data_points.append(pt1)
        pt2 = DataPoint(idx=2)
        pt2.graphicalProperties.solidFill = red
        s2.data_points.append(pt2)
    ds.add_chart(chart2, f"D{row - 3}")

    # ── Accessibility Summary ──
    row = qa_chart_row + 3
    ds.cell(row=row, column=1, value="ACCESSIBILITY (WCAG 2.1 AA)").font = section_font
    row += 1
    for label, count, color in [
        ("Critical", a11y_summary.get("Critical", 0), red),
        ("Warning", a11y_summary.get("Warning", 0), amber),
        ("Info", a11y_summary.get("Info", 0), "6A7883"),
    ]:
        ds.cell(row=row, column=1, value=label).font = label_font
        c = ds.cell(row=row, column=2, value=count)
        c.font = Font(name="Arial", size=14, bold=True, color=color)
        c.alignment = Alignment(horizontal="center")
        row += 1

    # ── Course Readiness ──
    row += 1
    ds.cell(row=row, column=1, value="COURSE READINESS").font = section_font
    row += 1
    # Mirror the HTML readiness logic: derive status from categories if no explicit 'overall'
    if readiness.get("overall"):
        readiness_status = readiness["overall"]
    elif readiness_cats:
        readiness_status = "READY" if readiness_pass == readiness_total else "NOT READY"
    else:
        readiness_status = "Not Evaluated"
    status_color = green if readiness_status == "READY" else (gray if readiness_status == "Not Evaluated" else red)
    ds.cell(row=row, column=1, value=f"Status: {readiness_status}").font = Font(
        name="Arial", size=16, bold=True, color=status_color
    )
    row += 1
    for cat in readiness_cats:
        icon = "✓" if cat.get("status") == "Pass" else "✗"
        color = green if cat.get("status") == "Pass" else red
        ds.cell(row=row, column=1, value=f"  {icon}  {cat.get('name', '')}").font = Font(
            name="Arial", size=10, color=color
        )
        row += 1

    # ── Confidence Distribution (Pie Chart) ──
    row += 1
    ds.cell(row=row, column=1, value="CONFIDENCE DISTRIBUTION").font = section_font
    row += 1

    # Count confidence across all standards items
    conf_counts = {"High": 0, "Medium": 0, "Low": 0}
    for item in sections.get("design_standards", {}).get("items", []):
        conf = item.get("confidence", "Medium")
        if conf in conf_counts:
            conf_counts[conf] += 1

    pie_data_row = row
    ds.cell(row=pie_data_row, column=1, value="Tier")
    ds.cell(row=pie_data_row, column=2, value="Count")
    for i, (tier, cnt) in enumerate(conf_counts.items()):
        ds.cell(row=pie_data_row + 1 + i, column=1, value=tier)
        ds.cell(row=pie_data_row + 1 + i, column=2, value=cnt)

    pie = PieChart()
    pie.title = "Confidence Distribution"
    pie.width = 14
    pie.height = 10
    pie_labels = Reference(ds, min_col=1, min_row=pie_data_row + 1, max_row=pie_data_row + 3)
    pie_data = Reference(ds, min_col=2, min_row=pie_data_row, max_row=pie_data_row + 3)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_labels)
    # Color slices
    if pie.series:
        s = pie.series[0]
        for idx, color in enumerate([green, amber, red]):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = color
            s.data_points.append(pt)
    ds.add_chart(pie, f"D{pie_data_row - 1}")

    # ── Action Items ──
    row = pie_data_row + 5
    ds.cell(row=row, column=1, value="TOP ACTION ITEMS").font = section_font
    row += 1

    # Gather not-met and low-confidence items
    action_items = []
    for item in sections.get("design_standards", {}).get("items", []):
        if item.get("status") == "Not Met" or item.get("confidence", "").lower() == "low":
            action_items.append(item)
    # Also add QA fails
    for item in sections.get("qa_categories", {}).get("items", []):
        if item.get("status") == "Fail":
            action_items.append({"id": item.get("id", ""), "name": item.get("name", ""), "status": "Fail", "recommendation": item.get("detail", "")})

    ds.cell(row=row, column=1, value="#").font = Font(name="Arial", size=10, bold=True)
    ds.cell(row=row, column=2, value="Standard/Category").font = Font(name="Arial", size=10, bold=True)
    ds.cell(row=row, column=3, value="Status").font = Font(name="Arial", size=10, bold=True)
    ds.cell(row=row, column=4, value="Recommendation").font = Font(name="Arial", size=10, bold=True)
    for col in range(1, 5):
        ds.cell(row=row, column=col).fill = PatternFill("solid", fgColor=gold)
    row += 1

    for i, ai in enumerate(action_items[:15], 1):
        ds.cell(row=row, column=1, value=i).font = label_font
        ds.cell(row=row, column=2, value=f"{ai.get('id', '')}. {ai.get('name', '')}").font = label_font
        status_val = ai.get("status", "")
        ds.cell(row=row, column=3, value=status_val).font = _status_font(status_val)
        ds.cell(row=row, column=4, value=ai.get("recommendation", ai.get("detail", ""))).font = label_font
        row += 1

    # ── Audit Scope Policy ──
    row += 2
    ds.merge_cells(f"A{row}:H{row}")
    ds.cell(row=row, column=1, value=(
        "Audit Scope: This audit evaluates content within Canvas. "
        "External links are listed in the External Links sheet but not reviewed unless external scanning is enabled. "
        '"Met" requires evidence across all relevant modules. "Partially Met" indicates evidence in some modules.'
    )).font = Font(name="Arial", size=9, italic=True, color="666666")

    # Column widths
    for col, width in [(1, 28), (2, 12), (3, 12), (4, 45), (5, 18), (6, 18), (7, 18), (8, 18)]:
        ds.column_dimensions[get_column_letter(col)].width = width


def _create_external_links_sheet(wb, data):
    """Create External Links inventory sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    el = wb.create_sheet("External Links")

    maroon = "8C1D40"
    gold = "FFC627"

    # Headers
    headers = ["Page", "Link Text", "URL", "Domain", "Status", "Notes"]
    for i, h in enumerate(headers, 1):
        cell = el.cell(row=1, column=i, value=h)
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=maroon)
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    for col, width in [(1, 25), (2, 30), (3, 50), (4, 20), (5, 16), (6, 35)]:
        el.column_dimensions[get_column_letter(col)].width = width

    # Populate from external_links in data
    ext_links = data.get("external_links", [])

    # Also extract from accessibility findings if no dedicated list
    if not ext_links:
        for item in data.get("sections", {}).get("accessibility", {}).get("items", []):
            if "link" in item.get("issue", "").lower():
                ext_links.append({
                    "page": item.get("page", ""),
                    "text": item.get("element", ""),
                    "url": "",
                    "domain": "",
                    "status": "Not Reviewed",
                    "notes": item.get("issue", ""),
                })

    row = 2
    for link in ext_links:
        el.cell(row=row, column=1, value=link.get("page", ""))
        el.cell(row=row, column=2, value=link.get("text", ""))
        url = link.get("url", "")
        cell_url = el.cell(row=row, column=3, value=url)
        if url:
            cell_url.font = Font(color="0563C1", underline="single")
            cell_url.hyperlink = url
        el.cell(row=row, column=4, value=link.get("domain", ""))
        status = link.get("status", "Not Reviewed")
        status_cell = el.cell(row=row, column=5, value=status)
        if status == "Reviewed":
            status_cell.font = Font(color="1E7E34")
        elif status == "Broken":
            status_cell.font = Font(color="C62828", bold=True)
        else:
            status_cell.font = Font(color="B5540A")
        el.cell(row=row, column=6, value=link.get("notes", ""))
        row += 1

    if not ext_links:
        el.cell(row=2, column=1, value="No external links detected").font = Font(italic=True, color="666666")
        row = 3

    # Policy footer
    row += 1
    el.merge_cells(f"A{row}:F{row}")
    el.cell(row=row, column=1, value=(
        "Audit Scope: This audit evaluates content within Canvas. "
        "External links are listed but not reviewed unless external scanning is enabled."
    )).font = Font(name="Arial", size=9, italic=True, color="666666")

    # Auto-filter on headers
    el.auto_filter.ref = f"A1:F{row - 2}"


# ============================================================
# FACULTY FEEDBACK SYNTHESIS
# ============================================================

# Maps standard IDs to plain-language labels for faculty communication
FACULTY_LABELS = {
    "01": "Learning Objectives Clarity",
    "02": "Module Objectives Alignment",
    "03": "Alignment Visibility for Students",
    "04": "Course Layout Consistency",
    "05": "Welcoming Course Introductions",
    "06": "Workload and Pacing Expectations",
    "07": "Instructor Guide for Future Faculty",
    "08": "Assessment-Objective Connection",
    "09": "Grading Transparency",
    "10": "Assessment Variety",
    "11": "Cognitive Skill Progression",
    "12": "Learning Materials-Objective Connection",
    "13": "Content Quality",
    "14": "Real-World Relevance",
    "15": "Inclusive and Diverse Content",
    "16": "Multiple Content Formats",
    "17": "Student Q&A Space",
    "18": "Instructor-Created Media",
    "19": "Active Learning Activities",
    "20": "Tool Integration",
    "21": "Technical and Academic Support",
    "22": "Content Accessibility (WCAG)",
    "23": "Third-Party Tool Accessibility",
    "24": "Mobile and Offline Access",
    "25": "Low-Cost Resources",
}


def generate_faculty_summary(data: dict) -> str:
    """Generate a plain-language faculty summary from audit results.

    Returns formatted plain text suitable for email or document sharing.
    Avoids jargon (CLO, MLO, WCAG, Bloom's) unless explained in context.

    Sections:
      1. Course overview and audit scope
      2. Findings summary (non-technical language)
      3. Action items grouped by priority (High / Medium / Low)
      4. What is already working well
    """
    course = data.get("course", {})
    course_name = course.get("name") or data.get("course_name", "the course")
    sections = data.get("sections", {})
    ds = sections.get("design_standards", {})
    ds_items = ds.get("items", [])
    ds_summary = ds.get("summary", {})

    lines = []
    lines.append("=" * 60)
    lines.append("COURSE QUALITY REVIEW — FACULTY SUMMARY")
    lines.append("=" * 60)
    lines.append("")
    lines.append(f"Course: {course_name}")
    lines.append(f"Review Date: {datetime.now().strftime('%B %d, %Y')}")
    lines.append(f"Reviewed by: ID Workbench (AI-assisted quality audit)")
    lines.append("")
    lines.append(
        "This summary translates the results of an automated quality review "
        "into plain language. It highlights what is working well, what needs "
        "attention, and specific action items for the course development team."
    )
    lines.append("")

    # ── What's Working Well ──
    met_items = [i for i in ds_items if i.get("status") == "Met"]
    lines.append("-" * 40)
    lines.append("WHAT'S WORKING WELL")
    lines.append("-" * 40)
    lines.append("")
    if met_items:
        for item in met_items:
            label = FACULTY_LABELS.get(item.get("id", ""), item.get("name", ""))
            lines.append(f"  [PASS]  {label}")
        lines.append("")
        lines.append(f"  {len(met_items)} of 25 quality standards are fully met.")
    else:
        lines.append("  No standards fully met yet — see action items below.")
    lines.append("")

    # ── Items Needing Attention ──
    high_priority = []
    medium_priority = []
    low_priority = []

    for item in ds_items:
        status = item.get("status", "")
        essential = item.get("essential", False)
        label = FACULTY_LABELS.get(item.get("id", ""), item.get("name", ""))
        rec = item.get("recommendation", "")
        evidence = item.get("evidence", "")

        if status == "Not Met":
            entry = {"label": label, "rec": rec, "evidence": evidence, "id": item.get("id")}
            if essential:
                high_priority.append(entry)
            else:
                medium_priority.append(entry)
        elif status == "Partially Met":
            entry = {"label": label, "rec": rec, "evidence": evidence, "id": item.get("id")}
            if essential:
                medium_priority.append(entry)
            else:
                low_priority.append(entry)

    if high_priority or medium_priority or low_priority:
        lines.append("-" * 40)
        lines.append("ACTION ITEMS")
        lines.append("-" * 40)
        lines.append("")

        if high_priority:
            lines.append("HIGH PRIORITY (required standards not currently met):")
            lines.append("")
            for entry in high_priority:
                lines.append(f"  * {entry['label']}")
                if entry["rec"]:
                    lines.append(f"    Action: {entry['rec']}")
                elif entry["evidence"]:
                    lines.append(f"    Current state: {entry['evidence'][:150]}")
                lines.append("")

        if medium_priority:
            lines.append("MEDIUM PRIORITY (partially met or needs improvement):")
            lines.append("")
            for entry in medium_priority:
                lines.append(f"  * {entry['label']}")
                if entry["rec"]:
                    lines.append(f"    Action: {entry['rec']}")
                elif entry["evidence"]:
                    lines.append(f"    Current state: {entry['evidence'][:150]}")
                lines.append("")

        if low_priority:
            lines.append("LOWER PRIORITY (nice-to-have improvements):")
            lines.append("")
            for entry in low_priority:
                lines.append(f"  * {entry['label']}")
                if entry["rec"]:
                    lines.append(f"    Suggestion: {entry['rec']}")
                lines.append("")
    else:
        lines.append("No action items — all standards are met!")
        lines.append("")

    # ── Summary Stats ──
    lines.append("-" * 40)
    lines.append("SUMMARY")
    lines.append("-" * 40)
    lines.append("")
    lines.append(f"  Standards Met:           {ds_summary.get('Met', 0)}")
    lines.append(f"  Partially Met:           {ds_summary.get('Partially Met', 0)}")
    lines.append(f"  Not Met:                 {ds_summary.get('Not Met', 0)}")
    na = ds_summary.get("Not Auditable", 0)
    if na:
        lines.append(f"  Not Auditable:           {na}")
    lines.append("")
    lines.append(
        "This summary was generated by ID Workbench. The instructional "
        "designer has reviewed the findings and will coordinate any "
        "necessary changes with you."
    )
    lines.append("")
    lines.append("=" * 60)

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description='Generate HTML or XLSX audit report')
    parser.add_argument('--input', type=str, help='Path to audit results JSON file')
    parser.add_argument('--demo', action='store_true', help='Generate demo report with sample data')
    parser.add_argument('--open', action='store_true', help='Open in browser/app after generating')
    parser.add_argument('--output', type=str, help='Custom output path')
    parser.add_argument('--xlsx', action='store_true', help='Generate XLSX report in QA Initiate format')
    parser.add_argument('--xlsx-output', type=str, help='Custom XLSX output path')
    parser.add_argument('--faculty', action='store_true', help='Generate plain-text faculty summary instead of HTML/XLSX report')
    args = parser.parse_args()

    if not args.demo:
        _track("skill_invoked", context={"skill": "audit"})

    # Load data
    if args.demo:
        data = generate_demo_data()
    elif args.input:
        data = json.loads(Path(args.input).read_text(encoding='utf-8'))
    else:
        if not sys.stdin.isatty():
            data = json.loads(sys.stdin.read())
        else:
            _log.error("Error: provide --input <file>, --demo, or pipe JSON to stdin")
            sys.exit(1)

    data = _normalize_audit_data(data)

    if args.faculty:
        summary = generate_faculty_summary(data)
        if args.output:
            out_path = Path(args.output)
        else:
            course_name, term, _ = _get_course_info(data)
            safe = _sanitize_folder_name(course_name)
            safe_term = _sanitize_folder_name(term)
            folder = _find_existing_course_folder(course_name, term) or (REPORTS_DIR / f"{safe}_{safe_term}")
            folder.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y-%m-%d")
            out_path = folder / f"faculty-summary_{timestamp}.txt"
        out_path.write_text(summary, encoding='utf-8')
        print(json.dumps({"ok": True, "path": str(out_path)}))
        if args.open:
            if sys.platform == 'darwin':
                subprocess.run(['open', str(out_path)])
            elif sys.platform == 'linux':
                subprocess.run(['xdg-open', str(out_path)])
        return

    if args.xlsx:
        # XLSX mode — save directly to reports/{course}/ (same as HTML).
        # Only use custom path if explicitly provided via --xlsx-output.
        archive_path = _build_report_path("xlsx", data)
        xlsx_path = args.xlsx_output or str(archive_path)
        result = generate_xlsx_report(data, xlsx_path)

        # Also save archive copy if custom path was used
        if xlsx_path != str(archive_path):
            try:
                import shutil
                shutil.copy2(xlsx_path, str(archive_path))
            except Exception as e:
                _log.warning(f"Could not save archive copy: {e}")

        result["archive_path"] = str(archive_path)

        # Upload XLSX to Supabase Storage (non-blocking)
        try:
            from metrics_sync import upload_report as _upload, is_configured
            if is_configured():
                _upload(str(archive_path))
        except Exception:
            pass

        # Sync audit score from XLSX results (non-blocking)
        try:
            if result.get("ok"):
                _total = result.get("met", 0) + result.get("partially_met", 0) + result.get("not_met", 0)
                if _total > 0:
                    _xlsx_score = round(result.get("met", 0) / _total * 100)
                    _cc, _term, _ = _get_course_info(data)
                    _cid = _cc if _cc != "UNKNOWN" else ""
                    if _cid:
                        from metrics_sync import sync_audit_score, is_configured as _is_cfg
                        if _is_cfg():
                            sync_audit_score(_cid, _xlsx_score)
        except Exception:
            pass

        # Push RLHF findings for XLSX-only mode
        try:
            rlhf_sid = push_to_rlhf(data, xlsx_path=str(archive_path))
            result["rlhf_session_id"] = rlhf_sid
        except Exception:
            pass

        print(json.dumps(result))

        if args.open and result.get("ok"):
            if sys.platform == 'darwin':
                subprocess.run(['open', result["path"]])
            elif sys.platform == 'linux':
                subprocess.run(['xdg-open', result["path"]])
    else:
        # HTML mode (default)
        # Audit reports are standalone deliverables — NOT staging files.
        # They save directly to reports/{course}/ with proper naming.
        html = generate_report(data)

        # 1. Save to reports/ with timestamped, course-scoped name
        archive_path = _build_report_path("html", data)
        archive_path.write_text(html, encoding='utf-8')
        _log.info(f"Audit report saved: {archive_path}")

        # 2. Also write to custom output if --output was specified
        if args.output:
            custom_path = Path(args.output)
            custom_path.parent.mkdir(parents=True, exist_ok=True)
            custom_path.write_text(html, encoding='utf-8')

        # 3. Upload to Supabase Storage (non-blocking)
        remote_url = None
        try:
            from metrics_sync import upload_report as _upload, is_configured
            if is_configured():
                remote_url = _upload(str(archive_path))
        except Exception:
            pass  # Never block on upload failure

        # 4. Sync audit score to Supabase (non-blocking)
        try:
            import re as _re
            score_match = _re.search(r'<div class="score-ring-value">(\d+)</div>', html)
            if score_match:
                _audit_score = int(score_match.group(1))
                _cc2, _t2, _ = _get_course_info(data)
                _course_id = _cc2 if _cc2 != "UNKNOWN" else ""
                if _course_id:
                    from metrics_sync import sync_audit_score, is_configured as _is_cfg
                    if _is_cfg():
                        sync_audit_score(_course_id, _audit_score)
        except Exception:
            pass  # Never block on score sync

        # 5. Push structured findings to RLHF Supabase (non-blocking)
        rlhf_session_id = None
        try:
            rlhf_session_id = push_to_rlhf(data, html_path=str(archive_path))
        except Exception:
            pass  # Never block on RLHF push

        result = {
            "ok": True,
            "path": str(archive_path),
            "archive_path": str(archive_path),
            "remote_url": remote_url,
            "rlhf_session_id": rlhf_session_id,
        }
        print(json.dumps(result))

        # Auto-open in browser (audit HTML is self-contained, no server needed)
        if args.open:
            if sys.platform == 'darwin':
                subprocess.run(['open', str(archive_path)])
            elif sys.platform == 'linux':
                subprocess.run(['xdg-open', str(archive_path)])


if __name__ == '__main__':
    main()
