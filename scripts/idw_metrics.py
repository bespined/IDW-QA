#!/usr/bin/env python3
"""Usage metrics tracker for ID Workbench.

Tracks skill invocations, pages built/pushed, audits run, fixes applied,
and time savings. Stores metrics in ~/.idw/metrics.json for ROI validation.

Usage in any script:
    from idw_metrics import track
    track("pages_built", count=3, context={"module": "Module 1"})
    track("audit_run", context={"mode": "standards", "findings": 12})

CLI usage:
    python idw_metrics.py --summary          # Print human-readable summary
    python idw_metrics.py --json             # Print raw JSON
    python idw_metrics.py --reset            # Clear all metrics
    python idw_metrics.py --since 2026-03-01 # Filter by date
"""

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

METRICS_DIR = Path.home() / ".idw"
METRICS_DIR.mkdir(parents=True, exist_ok=True)
METRICS_FILE = METRICS_DIR / "metrics.json"


def _load():
    """Load metrics from disk."""
    if METRICS_FILE.exists():
        try:
            return json.loads(METRICS_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return {"events": [], "totals": {}}
    return {"events": [], "totals": {}}


def _save(data):
    """Save metrics to disk."""
    METRICS_FILE.write_text(json.dumps(data, indent=2), encoding="utf-8")


def track(event_type, count=1, context=None):
    """Record a metric event.

    Args:
        event_type: Category of event. Standard types:
            - pages_built: Pages created via /course-build
            - pages_pushed: Pages pushed to Canvas
            - pages_rolled_back: Pages restored from backup
            - audit_run: QA/standards/a11y audit executed
            - audit_findings: Individual findings detected
            - audit_fixes: Auto-fixes applied via /bulk-edit
            - skill_invoked: Any skill invocation (concierge routes)
            - preview_generated: Unified preview or single preview
            - approval_exported: Approval manifest exported
            - report_generated: Audit HTML report created
            - api_calls: Canvas API calls made
            - backup_created: Pre-change backups saved
        count: Number of items (default 1)
        context: Optional dict with extra info (module name, skill name, etc.)
    """
    data = _load()

    # Auto-inject course_id from .env if not already in context
    if context is None:
        context = {}
    if "course_id" not in context:
        try:
            _env_path = Path(__file__).resolve().parents[1] / ".env"
            if _env_path.exists():
                for _line in _env_path.read_text().splitlines():
                    _line = _line.strip()
                    if _line.startswith("CANVAS_COURSE_ID="):
                        context["course_id"] = _line.split("=", 1)[1].strip()
                        break
        except Exception:
            pass

    event = {
        "type": event_type,
        "count": count,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }
    if context:
        event["context"] = context

    data["events"].append(event)

    # Update running totals
    if event_type not in data["totals"]:
        data["totals"][event_type] = 0
    data["totals"][event_type] += count

    _save(data)

    # Auto-sync to Supabase on significant events (non-blocking)
    _SYNC_EVENTS = {"pages_pushed", "audit_run", "report_generated", "pages_rolled_back"}
    if event_type in _SYNC_EVENTS:
        try:
            from metrics_sync import sync_metrics, is_configured
            if is_configured():
                sync_metrics(get_summary())
        except Exception:
            pass  # Never block on sync failure


def get_summary(since=None):
    """Generate a summary of all tracked metrics.

    Args:
        since: Optional ISO date string to filter events (e.g., "2026-03-01")

    Returns:
        Dict with totals, session_count, and time range.
    """
    data = _load()
    events = data.get("events", [])

    if since:
        events = [e for e in events if e.get("timestamp", "") >= since]

    # Recalculate totals from filtered events
    totals = {}
    for e in events:
        t = e["type"]
        totals[t] = totals.get(t, 0) + e.get("count", 1)

    # Calculate date range
    timestamps = [e.get("timestamp", "") for e in events if e.get("timestamp")]
    date_range = {
        "first": min(timestamps) if timestamps else None,
        "last": max(timestamps) if timestamps else None,
        "event_count": len(events),
    }

    # Derived ROI metrics — 20 min manual equivalent per action
    MANUAL_MINUTES_PER_ACTION = 20
    roi = {}
    pages = totals.get("pages_built", 0)
    if pages > 0:
        roi["pages_built"] = pages
        roi["estimated_hours_saved"] = round(pages * MANUAL_MINUTES_PER_ACTION / 60, 1)
    audits = totals.get("audit_run", 0)
    if audits > 0:
        roi["audits_run"] = audits
        roi["estimated_audit_hours_saved"] = round(audits * MANUAL_MINUTES_PER_ACTION / 60, 1)
    fixes = totals.get("audit_fixes", 0)
    if fixes > 0:
        roi["auto_fixes_applied"] = fixes
        roi["estimated_fix_hours_saved"] = round(fixes * MANUAL_MINUTES_PER_ACTION / 60, 1)

    return {
        "totals": totals,
        "events": events,
        "roi": roi,
        "date_range": date_range,
    }


def format_summary(summary):
    """Format summary dict as human-readable text."""
    lines = []
    lines.append("═══════════════════════════════════════════")
    lines.append(" ID Workbench — Usage Metrics")
    lines.append("═══════════════════════════════════════════")

    dr = summary["date_range"]
    if dr["first"]:
        lines.append(f" Period: {dr['first'][:10]} → {dr['last'][:10]}")
        lines.append(f" Total events: {dr['event_count']}")
    else:
        lines.append(" No events recorded yet.")
        return "\n".join(lines)

    lines.append("")
    lines.append(" Activity:")
    totals = summary["totals"]
    labels = {
        "pages_built": "  Pages built",
        "pages_pushed": "  Pages pushed to Canvas",
        "pages_rolled_back": "  Pages rolled back",
        "audit_run": "  Audits run",
        "audit_findings": "  Findings detected",
        "audit_fixes": "  Auto-fixes applied",
        "skill_invoked": "  Skill invocations",
        "preview_generated": "  Previews generated",
        "approval_exported": "  Approval manifests",
        "report_generated": "  Audit reports",
        "api_calls": "  Canvas API calls",
        "backup_created": "  Backups created",
    }
    for key, label in labels.items():
        if key in totals:
            lines.append(f"{label}: {totals[key]}")

    # Unlabeled totals
    for key in sorted(totals):
        if key not in labels:
            lines.append(f"  {key}: {totals[key]}")

    roi = summary.get("roi", {})
    if roi:
        lines.append("")
        lines.append(" Estimated Time Saved:")
        if "estimated_hours_saved" in roi:
            lines.append(f"  Page creation: {roi['estimated_hours_saved']} hrs")
        if "estimated_audit_hours_saved" in roi:
            lines.append(f"  QA auditing: {roi['estimated_audit_hours_saved']} hrs")
        if "estimated_fix_hours_saved" in roi:
            lines.append(f"  Auto-fixes: {roi['estimated_fix_hours_saved']} hrs")
        total_saved = sum(v for k, v in roi.items() if k.startswith("estimated_"))
        if total_saved > 0:
            lines.append(f"  ─────────────────────")
            lines.append(f"  Total: {total_saved:.1f} hrs saved")

    lines.append("═══════════════════════════════════════════")
    return "\n".join(lines)


ALL_SKILLS = [
    "course-build", "quiz", "assignment-generator", "discussion-generator",
    "rubric-creator", "audit", "bulk-edit", "course-transfer",
    "course-plan-import", "course-config", "canvas-nav", "media-upload",
    "update-module", "templates", "studio", "qti-import", "rubric-csv-import",
    "knowledge", "reorder-items", "notebooklm", "groups", "announcements",
    "analytics", "peer-review", "conferences", "blueprints", "mastery-paths",
    "outcomes", "accessibility",
]


def generate_dashboard(since=None, output_dir=None):
    """Generate an XLSX dashboard with 4 sheets: Raw Data, ROI, Engineering, Recommendations."""
    from collections import Counter
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = _load()
    events = data.get("events", [])
    if since:
        events = [e for e in events if e.get("timestamp", "") >= since]

    summary = get_summary(since=since)
    totals = summary["totals"]
    roi = summary.get("roi", {})
    dr = summary["date_range"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E2761", fill_type="solid")
    title_font = Font(bold=True, size=16)
    bold_font = Font(bold=True)

    wb = Workbook()

    # ── Sheet 1: Raw Data ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Raw Data"
    headers = ["Timestamp", "Event Type", "Count", "Context"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 8
    ws1.column_dimensions["D"].width = 50

    if events:
        for i, ev in enumerate(events, 2):
            ws1.cell(row=i, column=1, value=ev.get("timestamp", ""))
            ws1.cell(row=i, column=2, value=ev.get("type", ""))
            ws1.cell(row=i, column=3, value=ev.get("count", 1))
            ctx = ev.get("context")
            ws1.cell(row=i, column=4, value=json.dumps(ctx) if ctx else "")
        ws1.auto_filter.ref = f"A1:D{len(events) + 1}"
    else:
        ws1.cell(row=2, column=1, value="No data recorded.")
        ws1.auto_filter.ref = "A1:D2"

    # ── Sheet 2: ROI Dashboard ─────────────────────────────────────────
    ws2 = wb.create_sheet("ROI Dashboard")
    ws2.merge_cells("A1:D1")
    ws2.cell(row=1, column=1, value="ID Workbench — ROI Dashboard").font = title_font
    date_label = f"{dr['first'][:10]} to {dr['last'][:10]}" if dr.get("first") else "No data"
    ws2.cell(row=2, column=1, value=f"Date range: {date_label}").font = bold_font

    # Activity summary table
    ws2.cell(row=4, column=1, value="Activity").font = bold_font
    ws2.cell(row=4, column=2, value="Count").font = bold_font
    for c in (ws2.cell(row=4, column=1), ws2.cell(row=4, column=2)):
        c.fill = header_fill
        c.font = header_font
    activity_labels = {
        "pages_built": "Pages Built",
        "pages_pushed": "Pages Pushed",
        "pages_rolled_back": "Pages Rolled Back",
        "audit_run": "Audits Run",
        "audit_findings": "Findings Detected",
        "audit_fixes": "Auto-Fixes Applied",
        "skill_invoked": "Skill Invocations",
        "preview_generated": "Previews Generated",
        "api_calls": "API Calls",
        "backup_created": "Backups Created",
    }
    row = 5
    for key, label in activity_labels.items():
        ws2.cell(row=row, column=1, value=label)
        ws2.cell(row=row, column=2, value=totals.get(key, 0))
        row += 1
    activity_end = row - 1

    # Bar chart – activity counts
    bar = BarChart()
    bar.type = "col"
    bar.title = "Activity Counts by Event Type"
    bar.y_axis.title = "Count"
    bar.x_axis.title = "Activity"
    bar_data = Reference(ws2, min_col=2, min_row=4, max_row=activity_end)
    bar_cats = Reference(ws2, min_col=1, min_row=5, max_row=activity_end)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_cats)
    bar.width = 18
    bar.height = 12
    ws2.add_chart(bar, "E8")

    # ROI table
    roi_row = activity_end + 2
    ws2.cell(row=roi_row, column=1, value="Time Savings Category").font = bold_font
    ws2.cell(row=roi_row, column=2, value="Hours Saved").font = bold_font
    for c in (ws2.cell(row=roi_row, column=1), ws2.cell(row=roi_row, column=2)):
        c.fill = header_fill
        c.font = header_font
    roi_items = [
        ("Page creation", roi.get("estimated_hours_saved", 0)),
        ("QA auditing", roi.get("estimated_audit_hours_saved", 0)),
        ("Auto-fixes", roi.get("estimated_fix_hours_saved", 0)),
    ]
    r = roi_row + 1
    for label, val in roi_items:
        ws2.cell(row=r, column=1, value=label)
        ws2.cell(row=r, column=2, value=val)
        r += 1
    total_saved = sum(v for _, v in roi_items)
    ws2.cell(row=r, column=1, value="Total").font = bold_font
    ws2.cell(row=r, column=2, value=round(total_saved, 1)).font = bold_font
    roi_end = r

    # Pie chart – time savings breakdown
    pie = PieChart()
    pie.title = "Time Savings Breakdown"
    pie_data = Reference(ws2, min_col=2, min_row=roi_row, max_row=roi_end - 1)
    pie_cats = Reference(ws2, min_col=1, min_row=roi_row + 1, max_row=roi_end - 1)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    pie.width = 12
    pie.height = 10
    ws2.add_chart(pie, "E22")

    # ── Sheet 3: Engineering Dashboard ─────────────────────────────────
    ws3 = wb.create_sheet("Engineering Dashboard")
    ws3.merge_cells("A1:D1")
    ws3.cell(row=1, column=1, value="ID Workbench — Engineering Dashboard").font = title_font

    # Skill Usage
    skill_counts = Counter()
    for ev in events:
        if ev.get("type") == "skill_invoked":
            ctx = ev.get("context", {})
            if isinstance(ctx, dict) and "skill" in ctx:
                skill_counts[ctx["skill"]] += ev.get("count", 1)
    row = 3
    ws3.cell(row=row, column=1, value="Skill Usage").font = bold_font
    row += 1
    ws3.cell(row=row, column=1, value="Skill").font = header_font
    ws3.cell(row=row, column=1).fill = header_fill
    ws3.cell(row=row, column=2, value="Invocations").font = header_font
    ws3.cell(row=row, column=2).fill = header_fill
    skill_header_row = row
    row += 1
    for skill, cnt in sorted(skill_counts.items(), key=lambda x: -x[1]):
        ws3.cell(row=row, column=1, value=skill)
        ws3.cell(row=row, column=2, value=cnt)
        row += 1
    skill_end = row - 1

    if skill_counts:
        skill_bar = BarChart()
        skill_bar.type = "bar"
        skill_bar.title = "Skill Invocations"
        skill_bar.y_axis.title = "Skill"
        skill_bar.x_axis.title = "Count"
        s_data = Reference(ws3, min_col=2, min_row=skill_header_row, max_row=skill_end)
        s_cats = Reference(ws3, min_col=1, min_row=skill_header_row + 1, max_row=skill_end)
        skill_bar.add_data(s_data, titles_from_data=True)
        skill_bar.set_categories(s_cats)
        skill_bar.width = 16
        skill_bar.height = 10
        ws3.add_chart(skill_bar, "D3")

    # Error Rate
    row += 1
    ws3.cell(row=row, column=1, value="Error Rate by Script").font = bold_font
    row += 1
    error_counts = Counter()
    for ev in events:
        if ev.get("type") == "error_occurred":
            ctx = ev.get("context", {})
            script = ctx.get("script", "unknown") if isinstance(ctx, dict) else "unknown"
            error_counts[script] += ev.get("count", 1)
    ws3.cell(row=row, column=1, value="Script").font = header_font
    ws3.cell(row=row, column=1).fill = header_fill
    ws3.cell(row=row, column=2, value="Errors").font = header_font
    ws3.cell(row=row, column=2).fill = header_fill
    err_header_row = row
    row += 1
    for script, cnt in sorted(error_counts.items(), key=lambda x: -x[1]):
        ws3.cell(row=row, column=1, value=script)
        ws3.cell(row=row, column=2, value=cnt)
        row += 1
    err_end = row - 1

    if error_counts:
        err_bar = BarChart()
        err_bar.type = "col"
        err_bar.title = "Errors by Script"
        e_data = Reference(ws3, min_col=2, min_row=err_header_row, max_row=err_end)
        e_cats = Reference(ws3, min_col=1, min_row=err_header_row + 1, max_row=err_end)
        err_bar.add_data(e_data, titles_from_data=True)
        err_bar.set_categories(e_cats)
        err_bar.width = 14
        err_bar.height = 10
        ws3.add_chart(err_bar, f"D{err_header_row}")

    # API Usage by Hour
    row += 1
    ws3.cell(row=row, column=1, value="API Usage by Hour").font = bold_font
    row += 1
    hour_counts = Counter()
    for ev in events:
        if ev.get("type") == "api_calls":
            ts = ev.get("timestamp", "")
            try:
                hour = datetime.fromisoformat(ts).hour
                hour_counts[hour] += ev.get("count", 1)
            except (ValueError, TypeError):
                pass
    green_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
    ws3.cell(row=row, column=1, value="Hour").font = header_font
    ws3.cell(row=row, column=1).fill = header_fill
    ws3.cell(row=row, column=2, value="API Calls").font = header_font
    ws3.cell(row=row, column=2).fill = header_fill
    row += 1
    max_hour_val = max(hour_counts.values()) if hour_counts else 1
    for h in range(24):
        ws3.cell(row=row, column=1, value=h)
        c = ws3.cell(row=row, column=2, value=hour_counts.get(h, 0))
        if max_hour_val > 0:
            ratio = hour_counts.get(h, 0) / max_hour_val
            c.fill = red_fill if ratio > 0.5 else green_fill
        row += 1

    # Rollback Rate
    row += 1
    ws3.cell(row=row, column=1, value="Rollback Rate").font = bold_font
    row += 1
    pushed = totals.get("pages_pushed", 0)
    rolled_back = totals.get("pages_rolled_back", 0)
    success_pct = round((1 - rolled_back / pushed) * 100, 1) if pushed > 0 else 100.0
    ws3.cell(row=row, column=1, value="Pages Pushed")
    ws3.cell(row=row, column=2, value=pushed)
    row += 1
    ws3.cell(row=row, column=1, value="Pages Rolled Back")
    ws3.cell(row=row, column=2, value=rolled_back)
    row += 1
    rate_cell = ws3.cell(row=row, column=1, value=f"Success Rate: {success_pct}%")
    rate_cell.font = bold_font
    if success_pct < 95:
        rate_cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
    else:
        rate_cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")

    # Audit Finding Patterns
    row += 2
    ws3.cell(row=row, column=1, value="Audit Finding Patterns").font = bold_font
    row += 1
    finding_counts = Counter()
    for ev in events:
        if ev.get("type") == "audit_findings":
            ctx = ev.get("context", {})
            ftype = ctx.get("finding_type", "unknown") if isinstance(ctx, dict) else "unknown"
            finding_counts[ftype] += ev.get("count", 1)
    ws3.cell(row=row, column=1, value="Finding Type").font = header_font
    ws3.cell(row=row, column=1).fill = header_fill
    ws3.cell(row=row, column=2, value="Count").font = header_font
    ws3.cell(row=row, column=2).fill = header_fill
    find_header_row = row
    row += 1
    for ftype, cnt in sorted(finding_counts.items(), key=lambda x: -x[1]):
        ws3.cell(row=row, column=1, value=ftype)
        ws3.cell(row=row, column=2, value=cnt)
        row += 1
    find_end = row - 1

    if finding_counts:
        find_bar = BarChart()
        find_bar.type = "col"
        find_bar.title = "Top Recurring Finding Types"
        f_data = Reference(ws3, min_col=2, min_row=find_header_row, max_row=find_end)
        f_cats = Reference(ws3, min_col=1, min_row=find_header_row + 1, max_row=find_end)
        find_bar.add_data(f_data, titles_from_data=True)
        find_bar.set_categories(f_cats)
        find_bar.width = 14
        find_bar.height = 10
        ws3.add_chart(find_bar, f"D{find_header_row}")

    # Build Completion
    row += 1
    ws3.cell(row=row, column=1, value="Build Completion").font = bold_font
    row += 1
    build_status = Counter()
    for ev in events:
        if ev.get("type") == "build_phase":
            ctx = ev.get("context", {})
            status = ctx.get("status", "unknown") if isinstance(ctx, dict) else "unknown"
            build_status[status] += ev.get("count", 1)
    ws3.cell(row=row, column=1, value="Status").font = header_font
    ws3.cell(row=row, column=1).fill = header_fill
    ws3.cell(row=row, column=2, value="Count").font = header_font
    ws3.cell(row=row, column=2).fill = header_fill
    row += 1
    for status in ["started", "completed", "failed"]:
        ws3.cell(row=row, column=1, value=status)
        ws3.cell(row=row, column=2, value=build_status.get(status, 0))
        row += 1

    # ── Sheet 4: Recommendations ───────────────────────────────────────
    ws4 = wb.create_sheet("Recommendations")
    ws4.merge_cells("A1:D1")
    ws4.cell(row=1, column=1, value="ID Workbench — Recommendations").font = title_font
    ws4.column_dimensions["A"].width = 80

    row = 3
    recommendations = []

    # Skills with 0 invocations
    for skill in ALL_SKILLS:
        if skill_counts.get(skill, 0) == 0:
            recommendations.append(
                f"\u26a0\ufe0f {skill} has never been used. Consider introducing it to your workflow."
            )

    # Low adoption skills (< 3 uses in 30 days)
    for skill in ALL_SKILLS:
        cnt = skill_counts.get(skill, 0)
        if 0 < cnt < 3:
            recommendations.append(
                f"\U0001f4da {skill} has low adoption ({cnt} uses). Additional education or documentation may help."
            )

    # Rollback rate > 5%
    if pushed > 0:
        rollback_rate = rolled_back / pushed
        if rollback_rate > 0.05:
            recommendations.append(
                f"\u26a0\ufe0f Rollback rate is {rollback_rate * 100:.1f}%. "
                "Investigate common causes of rollbacks to improve first-push success."
            )

    # One finding type > 30% of all findings
    total_findings = sum(finding_counts.values()) if finding_counts else 0
    for ftype, cnt in finding_counts.items():
        if total_findings > 0 and cnt / total_findings > 0.30:
            recommendations.append(
                f"\U0001f527 '{ftype}' accounts for {cnt / total_findings * 100:.0f}% of audit findings. "
                "Consider automating a fix for this pattern."
            )

    # Error rate > 10% for a script
    total_events = len(events)
    for script, cnt in error_counts.items():
        if total_events > 0 and cnt / total_events > 0.10:
            recommendations.append(
                f"\U0001f6a8 {script} has a high error rate ({cnt} errors / {total_events} events = "
                f"{cnt / total_events * 100:.0f}%). Investigate root causes."
            )

    if not recommendations:
        recommendations.append("No actionable recommendations at this time. Keep up the good work!")

    for rec in recommendations:
        ws4.cell(row=row, column=1, value=rec)
        row += 1

    # Save file
    outdir = Path(output_dir) if output_dir else METRICS_DIR
    outdir.mkdir(parents=True, exist_ok=True)
    filename = f"idw_dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = outdir / filename
    wb.save(str(filepath))
    return str(filepath)


def main():
    parser = argparse.ArgumentParser(description="ID Workbench usage metrics")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--summary", action="store_true", help="Human-readable summary")
    group.add_argument("--json", action="store_true", help="Raw JSON output")
    group.add_argument("--reset", action="store_true", help="Clear all metrics")
    group.add_argument("--dashboard", action="store_true", help="Generate XLSX dashboard")
    group.add_argument("--track", type=str, metavar="EVENT_TYPE",
                       help="Record a metric event (e.g., skill_invoked)")

    parser.add_argument("--since", help="Filter events since date (YYYY-MM-DD)")
    parser.add_argument("--count", type=int, default=1, help="Event count (default: 1)")
    parser.add_argument("--context", type=str, default=None,
                        help='JSON context for --track (e.g., \'{"skill": "audit"}\')')
    parser.add_argument("-o", "--output", help="Output directory for dashboard (default: ~/.idw/)")
    args = parser.parse_args()

    if args.track:
        ctx = None
        if args.context:
            try:
                ctx = json.loads(args.context)
            except json.JSONDecodeError:
                print(f"Invalid JSON context: {args.context}", file=sys.stderr)
                sys.exit(1)
        track(args.track, count=args.count, context=ctx)
        print(json.dumps({"tracked": args.track, "count": args.count}))
        return

    if args.reset:
        _save({"events": [], "totals": {}})
        print("Metrics cleared.")
        return

    if args.dashboard:
        outdir = args.output or str(METRICS_DIR)
        path = generate_dashboard(since=args.since, output_dir=outdir)
        print(f"Dashboard saved to: {path}")
        return

    summary = get_summary(since=args.since)

    if args.json:
        # Exclude raw events from JSON output (too verbose) — use --dashboard for full data
        output = {k: v for k, v in summary.items() if k != "events"}
        print(json.dumps(output, indent=2))
    else:
        print(format_summary(summary))


if __name__ == "__main__":
    main()
